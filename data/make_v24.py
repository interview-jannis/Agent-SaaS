import openpyxl, shutil

shutil.copy2(
    'E:/Interview Co/10. Agent SaaS/Code/agent-saas/data/products_master_v23.xlsx',
    'E:/Interview Co/10. Agent SaaS/Code/agent-saas/data/products_master_v24.xlsx'
)

# Format: ★ for metadata headers, ▶ for sub-section headers, • for items.
REWRITES = {

# ═════════════════════════════════════════════════════════════════════════════
# K-Wellness · K-Shopping (P-208~210)
# ═════════════════════════════════════════════════════════════════════════════
'#P-208': (
    '★ Premium International Designer Eyewear\n'
    '★ Location : Seoul (multiple branches)\n'
    '★ Program\n'
    '   ▶ Specialty\n'
    '   • Korea\'s largest Lindberg dealer — premium international designer eyewear\n'
    '   • State-of-the-art digital lens fitting and advanced 3D diagnostic equipment\n'
    '   • Highly trained optometrists with precision fitting systems\n'
    '   ▶ Branches\n'
    '   • Dogok • Samsung Medical Center • Apgujeong • KINTEX • Pangyo • Cheonho\n'
    '   • Express Bus Terminal • Apgujeong Rodeo • Gwanggyo • Yeouido • Myeongdong • Jamsil'
),
'#P-209': (
    '★ K-Beauty Health & Beauty Retail\n'
    '★ Location : Seoul (Myeongdong Town, Gangnam Town, Seongsu N flagship)\n'
    '★ Program\n'
    '   ▶ Specialty\n'
    '   • Korea\'s first and largest health & beauty retail chain\n'
    '   • Curated selection of K-cosmetics, health supplements, beauty tools, and lifestyle products\n'
    '   • Mix of Korean indie brands and global favorites under one roof'
),
'#P-210': (
    '★ K-Stationery & Lifestyle Brand\n'
    '★ Location : Seoul (Myeongdong, multiple branches)\n'
    '★ Program\n'
    '   ▶ Specialty\n'
    '   • Korea\'s first design-focused stationery and lifestyle brand\n'
    '   • Pioneered the transformation of stationery into a design-driven lifestyle category\n'
    '   • Stylish stationery, planners, home goods, and K-culture merchandise'
),

# ═════════════════════════════════════════════════════════════════════════════
# K-Wellness · K-Content (P-211~224)
# ═════════════════════════════════════════════════════════════════════════════
'#P-211': (
    '★ Immersive XR/AR K-POP Adventure\n'
    '★ Duration : 1 day\n'
    '★ Min. 2 / Max. 6–7 persons per group\n'
    '★ Program\n'
    '   ▶ Route\n'
    '   • Gyeongbokgung → Bukchon/Seochon → Myeongdong\n'
    '   • → VAN Entertainment DIY Workshop → Naksan Park → souvenir shopping\n'
    '   ▶ Inclusions\n'
    '   • 1 van • English guide • Professional photographer • Lunch'
),
'#P-212': (
    '★ SBS Inkigayo Live Taping Package\n'
    '★ Duration : 1 day (every Sunday)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Live taping ticket for SBS Inkigayo at SBS Public Hall\n'
    '   • Walking tour with monthly-changing itinerary\n'
    '   • Transportation to SBS Public Hall\n'
    '   ▶ Notes\n'
    '   • Passport required; exclusive to foreign visitors'
),
'#P-213': (
    '★ K-Drama Filming Location Tour (Seoul)\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Featured Dramas\n'
    '   • Descendants of the Sun • Guardian: The Lonely and Great God • My Love From the Star\n'
    '   ▶ Route\n'
    '   • Songdo Dalkomm Coffee → Songdo Central Park → Deoksugung\n'
    '   • → Hakrim Coffee → N Seoul Tower\n'
    '   ▶ Inclusions\n'
    '   • Lunch at BBQ restaurant (Guardian filming site) • Private guide throughout'
),
'#P-214': (
    '★ K-Drama Filming Location Tour (Jeju)\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Featured Dramas\n'
    '   • Welcome to Samdal-ri • Extraordinary Attorney Woo • When Life Gives You Tangerines\n'
    '   ▶ Route A\n'
    '   • Gwaneumsa Temple • Secret Forest Andol-oreum • Gwangchigi Beach\n'
    '   • Seongsan Ilchulbong • Seongiwpean Bakery\n'
    '   ▶ Route B\n'
    '   • Dodubong Peak • Rainbow Coastal Road • Iho Taewooja Horse Lighthouses\n'
    '   • Myeongwol Elementary School • Changsindo\n'
    '   ▶ Inclusions\n'
    '   • Private guide-accompanied tour • Vehicle included\n'
    '   • Meals at guest\'s discretion'
),
'#P-215': (
    '★ Squid Game Tour (Incheon)\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Filming Location\n'
    '   • Visit Gyodong Elementary — a real Squid Game filming location\n'
    '   ▶ On-Screen Challenges\n'
    '   • Dalgona carving • Punch machine\n'
    '   ▶ Activities\n'
    '   • Retro "lunchbox" meal as seen on screen\n'
    '   • Sea-view monorail loop\n'
    '   • Archery Café at Wolmi Island\n'
    '   • Songdo Central Park mission game\n'
    '   ▶ Inclusions\n'
    '   • Hotel pickup/drop-off'
),
'#P-216': (
    '★ VIP K-POP Idol Experience\n'
    '★ Duration : 1 day (max. 4 participants per day)\n'
    '★ Program\n'
    '   ▶ Training\n'
    '   • K-POP idol dance training • K-POP idol vocal training\n'
    '   ▶ Styling & Shoot\n'
    '   • Idol concept makeup • Idol concept photo shoot\n'
    '   ▶ Inclusions\n'
    '   • Lunch (Korean cuisine)'
),
'#P-217': (
    '★ K-POP One Day Dance Lesson\n'
    '★ Location : Seoul\n'
    '★ Duration : 120 minutes\n'
    '★ Program\n'
    '   ▶ Flow\n'
    '   • Lesson counseling → Instructor assignment → Lesson start\n'
    '   ▶ Features\n'
    '   • Private K-POP one-day dance class tailored to group skill level and purpose\n'
    '   • Small-group format with professional choreography instructor'
),
'#P-218': (
    '★ K-Fashion Personal Shopping Tour\n'
    '★ Location : Seoul, Gangnam\n'
    '★ Duration : 120 minutes\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • 1-on-1 styling session with a professional Korean fashion stylist\n'
    '   • 6–7 curated outfit recommendations per hour from popular Korean fashion brands\n'
    '   ▶ Notes\n'
    '   • Final purchase decisions entirely at the guest\'s discretion'
),
'#P-219': (
    '★ K-Beauty Makeup & Hair Styling\n'
    '★ Location : Seoul, Gangnam Cheongdam-dong (Dosan Park area)\n'
    '★ Duration : 240 minutes\n'
    '★ Program\n'
    '   ▶ Package\n'
    '   • Makeup • Hair Styling • Photoshoot\n'
    '   ▶ Studio\n'
    '   • Professional K-Beauty styling at studios where Korean celebrities prepare for performances\n'
    '   • Near Supreme, Louis Vuitton, Stussy'
),
'#P-220': (
    '★ Master Style Consulting\n'
    '★ Duration : 300 minutes\n'
    '★ Program\n'
    '   ▶ I Style Consulting\n'
    '   • Professional image consulting — hair, makeup, personal color, and body type analysis\n'
    '   ▶ Fit Style Consulting\n'
    '   • Advanced body-type fit styling based on data from 15,000+ measurements\n'
    '   ▶ Inclusions\n'
    '   • In-person analysis • Practical shopping guidance\n'
    '   • Combined I Style + Fit Style package'
),
'#P-221': (
    '★ VIP Private Consulting\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Ultra-premium VIP offline service by prior inquiry only\n'
    '   • Comprehensive consulting: hair, makeup, fashion, and signature style design\n'
    '   • 1:1 customized analysis and style proposal\n'
    '   ▶ Notes\n'
    '   • Inquiries via email only'
),
'#P-222': (
    '★ VIP Offline Consulting\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • VIP private consulting by a team of specialist consultants\n'
    '   • Includes: Beauty, Fashion, Hair, and Makeup consulting\n'
    '   ▶ Notes\n'
    '   • KakaoTalk channel inquiries accepted'
),
'#P-223': (
    '★ Master Premium Personal Color Diagnosis\n'
    '★ Duration : 150 minutes\n'
    '★ Program\n'
    '   ▶ Color Diagnosis\n'
    '   • 25-type detailed seasonal tone color draping analysis\n'
    '   ▶ Recommendations\n'
    '   • Color cosmetics • Fashion style • Accessories\n'
    '   • Perfume • Glasses/lens • Nail recommendations\n'
    '   ▶ Additional\n'
    '   • Body structure analysis • Strategic fashion coaching • Signature image consulting'
),
'#P-224': (
    '★ Palette Premium Personal Color Diagnosis\n'
    '★ Duration : 150 minutes\n'
    '★ Program\n'
    '   ▶ Color Diagnosis\n'
    '   • In-depth personal color diagnosis after body color analysis\n'
    '   • Hair, jewelry, and lens color recommendations\n'
    '   ▶ Additional\n'
    '   • Face zone mini analysis • Fashion coaching\n'
    '   • Perfume, necklace, earring, ring, and watch recommendations'
),

# ═════════════════════════════════════════════════════════════════════════════
# K-Wellness · Tour (P-225~234)
# ═════════════════════════════════════════════════════════════════════════════
'#P-225': (
    '★ VIP Special Full-Day Seoul City Tour\n'
    '★ Location : Seoul\n'
    '★ Duration : 1 day (09:00–17:30)\n'
    '★ Min. 4 persons\n'
    '★ Program\n'
    '   ▶ Itinerary\n'
    '   • Hotel → City Hall / Gwanghwamun / Cheonggyecheon → Royal Guard Ceremony\n'
    '   • → Gyeongbokgung → National Folk Museum → Amethyst / Ginseng Center\n'
    '   • → N Seoul Tower → Lunch → Changdeokgung → Insadong\n'
    '   • → Han River Cruise → Hotel\n'
    '   ▶ Inclusions\n'
    '   • Private vehicle • Professional English guide • Lunch'
),
'#P-226': (
    '★ Seoul City Night Tour\n'
    '★ Location : Seoul\n'
    '★ Duration : 4 hours\n'
    '★ Program\n'
    '   ▶ Itinerary\n'
    '   • Bamdokkaebi (Night Goblin) night driving tour of Seoul illuminated landmarks\n'
    '   • Route: Gwanghwamun Square → Cheonggyecheon → DDP → N Seoul Tower\n'
    '   • → Han River Park → Banpo Moonbow Fountain → Itaewon\n'
    '   ▶ Inclusions\n'
    '   • English-speaking guide\n'
    '   ▶ Notes\n'
    '   • Exclusive to foreign visitors • Passport required'
),
'#P-227': (
    '★ Islamic Seoul City Tour\n'
    '★ Location : Seoul\n'
    '★ Duration : 4 hours\n'
    '★ Muslim-only private tour with halal meals and dedicated prayer time\n'
    '★ Program\n'
    '   ▶ Itinerary\n'
    '   • Hotel → Hanbok → Royal Guard Ceremony → Gyeongbokgung → N Seoul Tower\n'
    '   • → Itaewon Mosque → Halal Lunch → Prayer Time → Bukchon Village\n'
    '   • → Dongdaemun Market → Gwangjang Market → Myeongdong\n'
    '   ▶ Inclusions\n'
    '   • Private vehicle • Halal lunch • Entrance fees • Arabic/English guide'
),
'#P-228': (
    '★ VIP Private DMZ + NLL Tour\n'
    '★ Location : Imjingak / Ganghwa Island, Gyeonggi-do\n'
    '★ Duration : 1 day (07:00–17:30)\n'
    '★ Min. 4 persons\n'
    '★ Program\n'
    '   ▶ Itinerary\n'
    '   • Hotel → Imjingak → Unification Bridge → Passport check\n'
    '   • → DMZ Theater & Exhibition → 3rd Tunnel → Dora Observatory\n'
    '   • → Unification Village → Lunch → Ganghwa Island coastal fence\n'
    '   • → Ganghwa Peace Observatory → Hotel\n'
    '   ▶ Inclusions\n'
    '   • Private vehicle • Professional guide • Lunch'
),
'#P-229': (
    '★ Islamic Nami Island Tour\n'
    '★ Location : Nami Island, Gangwon-do\n'
    '★ Duration : 4 hours\n'
    '★ Muslim-only private tour\n'
    '★ Program\n'
    '   ▶ Itinerary\n'
    '   • Hotel → Nami Island → Halal lunch → Prayer time\n'
    '   • → Garden of Morning Calm → Hotel or Myeongdong\n'
    '   ▶ Inclusions\n'
    '   • Private vehicle • Halal lunch • Arabic/English guide'
),
'#P-230': (
    '★ Jeju Island Private Tour\n'
    '★ Location : Jeju Island\n'
    '★ Duration : 3 days\n'
    '★ Program\n'
    '   ▶ Accommodation\n'
    '   • Grand Hyatt Jeju\n'
    '   ▶ Inclusions\n'
    '   • Round-trip airfare\n'
    '   • Private vehicle + driver (Solati 4–10P / Combi Bus 11P+)\n'
    '   • Meals per itinerary (breakfast, lunch, one dinner daily)\n'
    '   • Admission: Pacific Marina Yacht, Hwansang Forest Gotjawal Park + foot bath\n'
    '   • Full-body massage\n'
    '   • Travel insurance\n'
    '   ▶ Notes\n'
    '   • Itinerary customizable'
),
'#P-231': (
    '★ Jeju Island Golf Tour\n'
    '★ Location : Jeju Island\n'
    '★ Duration : 3 days\n'
    '★ Program\n'
    '   ▶ Golf\n'
    '   • 3 rounds of 18-hole golf\n'
    '   • 2 rounds at Blackstone CC + 1 round at Elysian CC\n'
    '   • Both courses ranked in Asia Top 100\n'
    '   ▶ Accommodation\n'
    '   • 2 nights at Blackstone Villa Suite (private villa)\n'
    '   ▶ Inclusions\n'
    '   • Private vehicle\n'
    '   ▶ Notes\n'
    '   • Meals not included'
),
'#P-232': (
    '★ Gyeongbuk Tourpass\n'
    '★ Location : Gyeongsangbuk-do\n'
    '★ Duration : 24 / 48 / 72-hour options\n'
    '★ Program\n'
    '   ▶ Coverage\n'
    '   • Free-pass ticket for major paid attractions across Gyeongsangbuk-do\n'
    '   • Single barcode grants unlimited access to multiple attractions\n'
    '   ▶ Regions\n'
    '   • Gyeongju • Pohang • Ulleungdo • Partner cafés\n'
    '   ▶ Notes\n'
    '   • Various route options available'
),
'#P-233': (
    '★ Busan One-Night Luxury Cruise\n'
    '★ Location : Busan Port\n'
    '★ Duration : 2 days (every Saturday)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Panstar Bridge Tour\n'
    '   • Onboard dinner buffet\n'
    '   • Live performances and events\n'
    '   • Fireworks\n'
    '   • Breakfast on return\n'
    '   ▶ Notes\n'
    '   • Rotating routes'
),
'#P-234': (
    '★ Grand Tour of Korea by Rail Cruise (AURA)\n'
    '★ Location : Seoul → Suncheon → Yeosu → Busan → Gyeongju → Jeongdongjin → Donghae\n'
    '★ Duration : 3 days\n'
    '★ Program\n'
    '   ▶ Inclusions (All-inclusive)\n'
    '   • Train fare • Connecting buses\n'
    '   • Onboard accommodation\n'
    '   • All meals & snacks • Beverages\n'
    '   • Entrance & activity fees\n'
    '   • Onboard events'
),

# ═════════════════════════════════════════════════════════════════════════════
# K-Wellness · Leisure (P-235~249)
# ═════════════════════════════════════════════════════════════════════════════
'#P-235': (
    '★ MONA Yongpyong Ski Package\n'
    '★ Location : MONA Yongpyong Resort, Gangwon-do Pyeongchang\n'
    '★ Duration : 1–2 days\n'
    '★ Program\n'
    '   ▶ Resort\n'
    '   • Korea\'s largest ski resort\n'
    '   • 28 slopes • 14 lifts • 13 FIS-certified runs\n'
    '   ▶ Rental\n'
    '   • Ski / snowboard equipment and clothing\n'
    '   • Lift ticket included\n'
    '   ▶ Lesson\n'
    '   • Professional instructor (group or private) — separate pricing'
),
'#P-236': (
    '★ High 1 Resort Ski Package\n'
    '★ Location : High 1 Resort, Gangwon-do Jeongseon\n'
    '★ Duration : 1–2 days\n'
    '★ Program\n'
    '   ▶ Resort\n'
    '   • 15 FIS-certified slopes • 5 lifts • 3 gondolas\n'
    '   • Sleigh riding also available\n'
    '   ▶ Rental\n'
    '   • Ski / snowboard equipment and clothing\n'
    '   • Lift ticket included\n'
    '   ▶ Lesson\n'
    '   • Professional instructor available'
),
'#P-237': (
    '★ Vivaldi Park Ski Package\n'
    '★ Location : Vivaldi Park Resort, Gangwon-do Hongcheon\n'
    '★ Duration : 1–2 days\n'
    '★ Program\n'
    '   ▶ Resort\n'
    '   • 13 slopes • 10 lifts\n'
    '   ▶ Rental\n'
    '   • Ski / snowboard equipment and clothing\n'
    '   • Lift ticket included\n'
    '   ▶ Lesson\n'
    '   • Professional instructor available'
),
'#P-238': (
    '★ Konjiam Resort Ski Package\n'
    '★ Location : Konjiam Resort, Gyeonggi-do Gwangju\n'
    '★ Duration : 1–2 days\n'
    '★ Program\n'
    '   ▶ Resort\n'
    '   • 9 slopes • 5 lifts\n'
    '   • Closest major ski resort to Seoul\n'
    '   ▶ Rental\n'
    '   • Ski / snowboard equipment and clothing\n'
    '   • Lift ticket included\n'
    '   ▶ Lesson\n'
    '   • Professional instructor available'
),
'#P-239': (
    '★ Caribbean Bay Water Park\n'
    '★ Location : Everland Resort, Gyeonggi-do Yongin\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Attractions\n'
    '   • Caribbean-themed wave pool • Diving pool\n'
    '   • Surfing ride • Slides\n'
    '   ▶ Pricing\n'
    '   • Seasonal pricing (Middle / High / Gold season)'
),
'#P-240': (
    '★ Ocean World Water Park\n'
    '★ Location : Vivaldi Park Resort, Gangwon-do Hongcheon\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Attractions\n'
    '   • Ancient Egyptian theme\n'
    '   • Wave pool • Mega slides • Adventure zones\n'
    '   ▶ Pricing\n'
    '   • Seasonal pricing (Mid-Summer / Hot Summer)'
),
'#P-241': (
    '★ Cimer Spa & Water Park\n'
    '★ Location : Paradise City Resort, Incheon Jung-gu\n'
    '★ Duration : 6 hours\n'
    '★ Program\n'
    '   ▶ Concept\n'
    '   • European art-spa aesthetics • Pool party\n'
    '   • Korean-style sauna (jjimjilbang)\n'
    '   ▶ Facilities\n'
    '   • Slides • Infinity pool\n'
    '   ▶ Pricing\n'
    '   • Seasonal pricing'
),
'#P-242': (
    '★ Water Leisure Day Package\n'
    '★ Location : Gyeonggi-do Gapyeong\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Unlimited water park access\n'
    '   • 1-time jet boat ride\n'
    '   • Water skiing • Wakeboarding • Towable rides • Slides\n'
    '   • Unlimited BBQ'
),
'#P-243': (
    '★ Water Leisure Afternoon Package\n'
    '★ Location : Gyeonggi-do Gapyeong\n'
    '★ Duration : Afternoon (1 day)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Water park access\n'
    '   • 1-time jet boat ride\n'
    '   • Water skiing • Wakeboarding • Towable rides • Slides\n'
    '   • Unlimited BBQ'
),
'#P-244': (
    '★ Baseball Skybox (Kiwoom Heroes)\n'
    '★ Location : Gocheok Sky Dome, Seoul Guro-gu\n'
    '★ Duration : 4 hours\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • VIP Skybox at Gocheok Sky Dome\n'
    '   • Dedicated entrance via Central C Gate\n'
    '   • Private parking\n'
    '   • Mascot photo service'
),
'#P-245': (
    '★ Baseball Skybox (SSG Landers)\n'
    '★ Location : Incheon SSG Landers Field, Incheon Michuhol-gu\n'
    '★ Duration : 4 hours\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Premium private VIP skybox seating\n'
    '   • Food and beverages available in-seat\n'
    '   ▶ Pricing\n'
    '   • Weekday / weekend seasonal pricing'
),
'#P-246': (
    '★ Soccer Skybox (FC Seoul)\n'
    '★ Location : Seoul World Cup Stadium, Seoul Mapo-gu\n'
    '★ Duration : 3 hours\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Private VIP skybox with premium stadium views\n'
    '   • In-seat dining service'
),
'#P-247': (
    '★ Lotte World 1-Day Pass + Magic Pass\n'
    '★ Location : Seoul Songpa-gu\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • 1-Day Pass to Lotte World (indoor + outdoor)\n'
    '   • Magic Pass for priority attraction access\n'
    '   ▶ Options\n'
    '   • 5-ride or 7-ride Magic Pass\n'
    '   • No queuing at priority attractions'
),
'#P-248': (
    '★ Everland Dream Tour\n'
    '★ Location : Gyeonggi-do Yongin\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • VIP guided park experience with full priority access\n'
    '   • Dedicated parking • Front gate priority entry\n'
    '   • Private guide throughout the park\n'
    '   • Priority boarding at all major attractions\n'
    '   • Commemorative photo\n'
    '   • Restaurant meal\n'
    '   ▶ Optional\n'
    '   • Safari Special Tour (subject to availability)'
),
'#P-249': (
    '★ Seoul Land 1-Day Pass\n'
    '★ Location : Gyeonggi-do Gwacheon\n'
    '★ Duration : 1 day\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Full-day access to Seoul Land amusement park\n'
    '   • Classic Korean theme park with rides, attractions, and family entertainment'
),

# ═════════════════════════════════════════════════════════════════════════════
# K-Starcation · K-POP Camp (P-301~303)
# ═════════════════════════════════════════════════════════════════════════════
'#P-301': (
    '★ K-POP CAMP · Basic\n'
    '★ Location : Seoul\n'
    '★ Duration : 3 days\n'
    '★ Program\n'
    '   ▶ Training\n'
    '   • Group K-POP basic training (vocal & dance fundamentals)\n'
    '   ▶ Styling & Shoot\n'
    '   • K-POP idol hair & makeup styling\n'
    '   • Premium studio profile photo shoot\n'
    '   • Short-form (Reels/Shorts) challenge video production\n'
    '   ▶ Completion\n'
    '   • Official completion certificate'
),
'#P-302': (
    '★ K-POP CAMP · Gold\n'
    '★ Location : Seoul\n'
    '★ Duration : 5 days\n'
    '★ Program\n'
    '   ▶ Training\n'
    '   • Small-group intensive K-POP training (vocal, dance, stage performance)\n'
    '   ▶ Styling & Shoot\n'
    '   • Premium idol styling (stage costume fitting on request)\n'
    '   • Individual concept photo shoot (album-style)\n'
    '   ▶ Production\n'
    '   • Professional studio music recording session\n'
    '   • Highlight music video (MV) production\n'
    '   • Performance showcase at venue\n'
    '   ▶ Completion\n'
    '   • Official completion certificate'
),
'#P-303': (
    '★ K-POP CAMP · Platinum\n'
    '★ Location : Seoul\n'
    '★ Duration : 5 days\n'
    '★ Program\n'
    '   ▶ Training\n'
    '   • 1:1 master training with dedicated vocal/dance director\n'
    '   ▶ Styling & Production\n'
    '   • VVIP dedicated styling and custom stage design\n'
    '   • Solo original track production & recording\n'
    '   • Full-length solo music video (MV) production\n'
    '   • Private showcase performance\n'
    '   ▶ VIP Care\n'
    '   • Dedicated protocol & VIP care (prayer room / halal catering, etc.)\n'
    '   ▶ Bonus\n'
    '   • Entertainment agency audition visit'
),

# ═════════════════════════════════════════════════════════════════════════════
# K-Education · School Tour (P-401~403)
# ═════════════════════════════════════════════════════════════════════════════
'#P-401': (
    '★ CITY PACKAGE TOUR · SEOUL TOUR\n'
    '★ Location : Seoul\n'
    '★ Duration : 5 days\n'
    '★ Min. 30 / Max. 40 persons\n'
    '★ Theme : School Exchange, AI, Media, K-POP\n'
    '★ Program\n'
    '   ▶ Day 1\n'
    '   • Incheon Airport → G Valley Industrial Experience Center → Hiker Ground\n'
    '   ▶ Day 2\n'
    '   • School Exchange (AM/PM) → Namsan Cable Car → N Seoul Tower\n'
    '   ▶ Day 3\n'
    '   • Seoul Robotics & AI Science Center → MMCA Seoul → KBS Station\n'
    '   • → National Assembly Museum → Han River Cruise\n'
    '   ▶ Day 4\n'
    '   • National Museum of Korea → Myeongdong\n'
    '   • → Real K-POP Dance Experience → Hongdae Street\n'
    '   ▶ Day 5\n'
    '   • Gyeongbokgung Palace → National Palace Museum → Incheon Airport'
),
'#P-402': (
    '★ CITY + LOCAL PACKAGE TOUR · SEOUL, GYEONGGI TOUR\n'
    '★ Location : Seoul, Gyeonggi\n'
    '★ Duration : 5 days\n'
    '★ Min. 30 / Max. 40 persons\n'
    '★ Theme : University Campuses, Future Technologies, Media, Theme Parks\n'
    '★ Program\n'
    '   ▶ Day 1\n'
    '   • Incheon Airport → Seodaemun Prison History Museum → War Memorial of Korea\n'
    '   ▶ Day 2\n'
    '   • University Campus Tour → Hongdae Street → MBC World Theme Park → Han River Cruise\n'
    '   ▶ Day 3\n'
    '   • LG Discovery Lab → Seoul Future Lab → National Museum of Korea → Insadong\n'
    '   ▶ Day 4\n'
    '   • Wolmi Sea Train → Wolmi Theme Park → Songdo Central Park\n'
    '   • → National Museum of World Letters → G-Tower Observatory\n'
    '   ▶ Day 5\n'
    '   • Gimbap Making Experience → National Aviation Museum → Incheon Airport'
),
'#P-403': (
    '★ CITY + LOCAL PACKAGE TOUR\n'
    '★ Location : Seoul + Regional\n'
    '★ Duration : 5 days\n'
    '★ Min. 30 / Max. 40 persons\n'
    '★ Theme : K-POP, Traditional Culture, Theme Parks, Future Industries\n'
    '★ Program\n'
    '   ▶ Day 1\n'
    '   • Incheon Airport → Gwangjang Market / Dongdaemun Market\n'
    '   ▶ Day 2\n'
    '   • University Campus Tour → K-POP Dance Experience → Apgujeong K-Star Road\n'
    '   ▶ Day 3\n'
    '   • Korean Folk Village → Everland\n'
    '   ▶ Day 4\n'
    '   • Gyeongbokgung Palace → Hiker Ground → Namsan Cable Car & Seoul Tower\n'
    '   • → Myeongdong → Painters Show / Musical Chef\n'
    '   ▶ Day 5\n'
    '   • Hyundai Motorstudio Goyang → Heiri Art Village → Incheon Airport'
),

# ═════════════════════════════════════════════════════════════════════════════
# Subpackage · Concierge (P-501~504)
# ═════════════════════════════════════════════════════════════════════════════
'#P-501': (
    '★ Personnel Support Fees\n'
    '★ Program\n'
    '   ▶ Hourly Rates\n'
    '   • Protocol Manager — 50,000–100,000 KRW / hr\n'
    '   • Protocol Assistant — 30,000–50,000 KRW / hr\n'
    '   • Driver — 40,000–70,000 KRW / hr\n'
    '   • Simultaneous Interpreter — 150,000–300,000 KRW / hr'
),
'#P-502': (
    '★ Personal Protocol\n'
    '★ Program\n'
    '   ▶ Hourly Rate\n'
    '   • Personal VIP Attendance — 100,000–200,000 KRW / hr'
),
'#P-503': (
    '★ Corporate Concierge\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Full-scope corporate concierge service for VIP groups and delegations\n'
    '   ▶ Pricing\n'
    '   • Price on inquiry'
),
'#P-504': (
    '★ VIP Hospitality\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Premium VIP hospitality and protocol management service\n'
    '   ▶ Pricing\n'
    '   • Price on inquiry'
),

# ═════════════════════════════════════════════════════════════════════════════
# Subpackage · Security (P-505~507)
# ═════════════════════════════════════════════════════════════════════════════
'#P-505': (
    '★ VIP Protocol & Security\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • VIP protocol and security service\n'
    '   ▶ Pricing\n'
    '   • Price on inquiry'
),
'#P-506': (
    '★ Personal Protection / VIP Protocol Service\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Personal protection and VIP protocol service\n'
    '   ▶ Pricing\n'
    '   • Price on inquiry'
),
'#P-507': (
    '★ Security Service\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Professional security service\n'
    '   ▶ Pricing\n'
    '   • Price on inquiry'
),

# ═════════════════════════════════════════════════════════════════════════════
# Subpackage · Hotel (P-508~544)
# ═════════════════════════════════════════════════════════════════════════════
'#P-508': (
    '★ Shilla Seoul\n'
    '★ Layout\n'
    '   ▶ Area : 53 m²\n'
    '   ▶ Composition\n'
    '   • 1 bedroom • 1 bathroom\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults or up to 4 with children'
),
'#P-509': (
    '★ Shilla Seoul\n'
    '★ Layout\n'
    '   ▶ Area : 84 m²\n'
    '   ▶ Composition\n'
    '   • 1 bedroom • 1 bathroom • 2 toilets • Living room\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 4 adults or up to 5 with children'
),
'#P-510': (
    '★ Shilla Seoul\n'
    '★ Layout\n'
    '   ▶ Area : 159 m²\n'
    '   ▶ Composition\n'
    '   • 1 bedroom • 1 bathroom • Sauna • 2 toilets\n'
    '   • Living room • Office • Dining room\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 4 adults or up to 5 with children\n'
    '★ Notes\n'
    '   • Phone reservation only'
),
'#P-511': (
    '★ Shilla Seoul\n'
    '★ Layout\n'
    '   ▶ Area : 66 m²\n'
    '   ▶ Composition\n'
    '   • 1 bedroom • 1 bathroom • 2 toilets\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 4 adults or up to 5 with children\n'
    '★ Notes\n'
    '   • Phone reservation only'
),
'#P-512': (
    '★ Signiel Seoul\n'
    '★ Area : 59–73 m²\n'
    '★ Occupancy : 2 adults\n'
    '★ View : Han River (view bathroom)'
),
'#P-513': (
    '★ Signiel Seoul\n'
    '★ Area : 70–86 m²\n'
    '★ Occupancy : 2 adults\n'
    '★ View : Han River (view bathroom)'
),
'#P-514': (
    '★ Signiel Seoul\n'
    '★ Area : 143 m²\n'
    '★ Occupancy : 2 adults\n'
    '★ View : City (view bathroom)'
),
'#P-515': (
    '★ Signiel Seoul\n'
    '★ Area : 84 m²\n'
    '★ Occupancy : 4 adults\n'
    '★ View : Han River'
),
'#P-516': (
    '★ Four Seasons Seoul\n'
    '★ Area : 48 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults or 1 adult + 2 infants\n'
    '★ View : Downtown or mountain view'
),
'#P-517': (
    '★ Four Seasons Seoul\n'
    '★ Area : 143 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults or 2 adults + 2 children\n'
    '★ View : Gyeongbokgung Palace view'
),
'#P-518': (
    '★ Park Hyatt Seoul\n'
    '★ Area : 52 m²'
),
'#P-519': (
    '★ Park Hyatt Seoul\n'
    '★ Area : 70 m²'
),
'#P-520': (
    '★ InterContinental Seoul\n'
    '★ Area : 50 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 4 adults'
),
'#P-521': (
    '★ InterContinental Seoul\n'
    '★ Area : 85 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults'
),
'#P-522': (
    '★ InterContinental Seoul\n'
    '★ Area : 120 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults'
),
'#P-523': (
    '★ Paradise City\n'
    '★ Area : 135 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults'
),
'#P-524': (
    '★ Paradise City\n'
    '★ Area : 84 m²\n'
    '★ Occupancy : 2 adults'
),
'#P-525': (
    '★ Paradise City\n'
    '★ Area : 942 m²\n'
    '★ Occupancy\n'
    '   • Standard : 4 adults\n'
    '   • Max : 5 adults'
),
'#P-526': (
    '★ Inspire Entertainment Resort\n'
    '★ Area : 131 m²\n'
    '★ Occupancy : 2 adults\n'
    '★ View : Lake view\n'
    '★ Inclusions\n'
    '   ▶ Dining\n'
    '   • Minagi Omakase set\n'
    '   ▶ Wellness\n'
    '   • Wellness Club access\n'
    '   • Complimentary indoor pool & fitness center access'
),
'#P-527': (
    '★ Inspire Entertainment Resort\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Back & lower back hair removal — 5-session course'
),
'#P-528': (
    '★ Inspire Entertainment Resort\n'
    '★ Area : 131 m²\n'
    '★ Occupancy : 2 adults\n'
    '★ View : Lake view\n'
    '★ Inclusions\n'
    '   ▶ Stay\n'
    '   • Late checkout\n'
    '   ▶ Splash Bay\n'
    '   • Afternoon pass • Cabana\n'
    '   ▶ Dining\n'
    '   • Chef\'s Kitchen dinner buffet\n'
    '   ▶ Wellness\n'
    '   • Wellness Club access\n'
    '   • Complimentary indoor pool & fitness center access'
),
'#P-529': (
    '★ Inspire Entertainment Resort\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Bikini line hair removal — 5-session course'
),
'#P-530': (
    '★ Park Hyatt Incheon\n'
    '★ Area : 108 m²'
),
'#P-531': (
    '★ Park Hyatt Incheon\n'
    '★ Area : 252 m²'
),
'#P-532': (
    '★ Oakwood Premier Incheon\n'
    '★ Layout\n'
    '   ▶ Area : 234 m²\n'
    '   ▶ Composition\n'
    '   • 3 bedrooms • Family room • 3 toilets\n'
    '   • Dining room • Kitchen • Office • Utility room\n'
    '★ Occupancy\n'
    '   • Standard : 8 adults\n'
    '   • Max : 9 adults'
),
'#P-533': (
    '★ Oakwood Premier Incheon\n'
    '★ Layout\n'
    '   ▶ Area : 371 m²\n'
    '   ▶ Composition\n'
    '   • 2 bedrooms • Family room • 2 toilets\n'
    '   • Dining room • 2 kitchens • Walk-in closet • Utility room\n'
    '★ Occupancy\n'
    '   • Standard : 4 adults\n'
    '   • Max : 5 adults'
),
'#P-534': (
    '★ Sheraton\n'
    '★ Area : 120 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults\n'
    '★ View : City view\n'
    '★ Inclusions\n'
    '   ▶ Club Lounge\n'
    '   • Club Lounge access\n'
    '   • Complimentary breakfast, snacks, hors d\'oeuvres\n'
    '   ▶ Business\n'
    '   • Complimentary business services'
),
'#P-535': (
    '★ Hanok Essay Hahoe\n'
    '★ Concept : Traditional Korean hanok accommodation\n'
    '★ Occupancy\n'
    '   • Standard : 4 adults\n'
    '   • Max : 6 adults\n'
    '★ Inclusions\n'
    '   ▶ Welcome Set\n'
    '   • Tea ceremony set with tea leaves\n'
    '   • Drip coffee set with beans'
),
'#P-536': (
    '★ Gyeongwonjae Ambassador\n'
    '★ Area : 132 m²\n'
    '★ Occupancy\n'
    '   • Standard : 2 adults\n'
    '   • Max : 3 adults'
),
'#P-537': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Area : 34 m²\n'
    '★ Bed : King or Twin'
),
'#P-538': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Area : 70 m²\n'
    '★ Bed : King\n'
    '★ Inclusions\n'
    '   • Executive Lounge access for 2 guests'
),
'#P-539': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Area : 135 m²\n'
    '★ Bed : Super King\n'
    '★ Inclusions\n'
    '   • Executive Lounge access for 2 guests'
),
'#P-540': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Area : 244 m²\n'
    '★ Bed : 2 Super King beds\n'
    '★ Inclusions\n'
    '   • Executive Lounge access for 2 guests\n'
    '★ Notes\n'
    '   • Special rate available — contact Sales Manager for inquiry'
),
'#P-541': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Area : 71 m² Residence\n'
    '★ Bed : King'
),
'#P-542': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Package : Noor Royal Indulgence Collection · Royal Arabian Indulgence\n'
    '★ Inclusions\n'
    '   ▶ Stay\n'
    '   • Guest room\n'
    '   ▶ Dining\n'
    '   • In-room dining (3-course menu)\n'
    '   ▶ Complimentary Gift\n'
    '   • Yoo Jian-jak handcrafted najeon (mother-of-pearl) gift box'
),
'#P-543': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Package : Serenity Sanctuary Escape · Safa Serenity Retreat\n'
    '★ Inclusions\n'
    '   ▶ Stay\n'
    '   • Guest room\n'
    '   ▶ Spa & Wellness\n'
    '   • Premium Spa & Wellness experience\n'
    '   • Biologique Recherche, Thermes Marins by certified therapists\n'
    '   ▶ Complimentary Gift\n'
    '   • Yoo Jian-jak handcrafted najeon gift box'
),
'#P-544': (
    '★ Sofitel Ambassador Seoul\n'
    '★ Package : K-Beauty Prestige Escape\n'
    '★ Inclusions\n'
    '   ▶ Stay\n'
    '   • Guest room\n'
    '   ▶ K-Beauty Experience\n'
    '   • Exclusive K-Beauty experience with premium devices\n'
    '   • Customized consulting • Gift set\n'
    '   ▶ Complimentary Gift\n'
    '   • Yoo Jian-jak handcrafted najeon gift box'
),

# ═════════════════════════════════════════════════════════════════════════════
# Subpackage · Interpreter (P-545~548)
# ═════════════════════════════════════════════════════════════════════════════
'#P-545': (
    '★ Arabic-Korean Interpretation\n'
    '★ Location : Seoul (available nationwide)\n'
    '★ Duration : 1 hour\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Professional Arabic-Korean simultaneous and consecutive interpretation\n'
    '   ▶ Specialization\n'
    '   • Medical settings • Business • VIP tourism'
),
'#P-546': (
    '★ Arabic-Korean Interpretation\n'
    '★ Location : Seoul (available nationwide)\n'
    '★ Duration : 6 hours\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Professional Arabic-Korean simultaneous and consecutive interpretation\n'
    '   ▶ Specialization\n'
    '   • Medical settings • Business • VIP tourism'
),
'#P-547': (
    '★ Arabic-Korean Interpretation\n'
    '★ Duration : 8 hours (daily rate)\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Professional Arabic-Korean interpretation\n'
    '   • General or Senior/Professional level\n'
    '   ▶ Specialization\n'
    '   • VIP clients • Medical consultations • High-profile engagements\n'
    '   ▶ Notes\n'
    '   • Overtime billed separately'
),
'#P-548': (
    '★ Arabic-Korean Interpretation\n'
    '★ Duration : 8 hours (daily rate)\n'
    '★ Program\n'
    '   ▶ Service\n'
    '   • Professional Arabic-Korean interpretation\n'
    '   • General or Senior/Professional level\n'
    '   ▶ Specialization\n'
    '   • VIP clients • Medical consultations • High-profile engagements\n'
    '   ▶ Notes\n'
    '   • Overtime billed separately'
),

# ═════════════════════════════════════════════════════════════════════════════
# Subpackage · Vehicle (P-549~562)
# ═════════════════════════════════════════════════════════════════════════════
'#P-549': (
    '★ Halal-Certified VIP Limousine\n'
    '★ Vehicle : Noble Klasse Solati S11\n'
    '★ Program\n'
    '   ▶ Specialty\n'
    '   • World\'s first halal-certified limousine in Korea\n'
    '   • Dedicated in-vehicle prayer space with qibla direction\n'
    '   • Prayer mat provided\n'
    '   ▶ Use Cases\n'
    '   • Muslim VIP-optimized mobility for travel, sightseeing, and medical visits\n'
    '   ▶ Pricing\n'
    '   • Price on inquiry'
),
'#P-550': (
    '★ VIP Charter Service\n'
    '★ Duration : 8 hours (base fare)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Options\n'
    '   • English-speaking driver available on request (additional charge)'
),
'#P-551': (
    '★ VIP Charter Service\n'
    '★ Duration : 8 hours (base fare)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Options\n'
    '   • English-speaking driver available on request (additional charge)'
),
'#P-552': (
    '★ VIP Charter Service\n'
    '★ Duration : 8 hours (base fare)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Options\n'
    '   • English-speaking driver available on request (additional charge)'
),
'#P-553': (
    '★ VIP Charter Service\n'
    '★ Duration : 8 hours (base fare)\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Options\n'
    '   • English-speaking driver available on request (additional charge)'
),
'#P-554': (
    '★ VIP Charter Service\n'
    '★ Duration : 10 hours\n'
    '★ Fleet : Genesis G90 RS4 / Mercedes S-Class / Mercedes Sprinter VIP\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   • All three vehicle types priced equally\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation'
),
'#P-555': (
    '★ VIP Charter Service\n'
    '★ Duration : 10 hours\n'
    '★ Fleet : Genesis G90 RS4 / Mercedes S-Class / Mercedes Sprinter VIP\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   • All three vehicle types priced equally\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation'
),
'#P-556': (
    '★ VIP Charter Service\n'
    '★ Duration : 10 hours\n'
    '★ Fleet : Genesis G90 RS4 / Mercedes S-Class / Mercedes Sprinter VIP\n'
    '★ Program\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   • All three vehicle types priced equally\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation'
),
'#P-557': (
    '★ VIP Charter Service\n'
    '★ Duration : 9 hours (09:00–18:00)\n'
    '★ Program\n'
    '   ▶ Eligible Clients\n'
    '   • Foreign VIPs/buyers • Government bodies • Corporate accounts\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Notes\n'
    '   • Overtime billed per hour after 18:00'
),
'#P-558': (
    '★ VIP Charter Service\n'
    '★ Duration : 9 hours (09:00–18:00)\n'
    '★ Program\n'
    '   ▶ Eligible Clients\n'
    '   • Foreign VIPs/buyers • Government bodies • Corporate accounts\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Notes\n'
    '   • Overtime billed per hour after 18:00'
),
'#P-559': (
    '★ VIP Charter Service\n'
    '★ Duration : 9 hours (09:00–18:00)\n'
    '★ Program\n'
    '   ▶ Eligible Clients\n'
    '   • Foreign VIPs/buyers • Government bodies • Corporate accounts\n'
    '   ▶ Inclusions\n'
    '   • Vehicle • Driver\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Parking • Driver meals/accommodation\n'
    '   ▶ Notes\n'
    '   • Overtime billed per hour after 18:00'
),
'#P-560': (
    '★ VIP Charter Service\n'
    '★ Vehicle : Premium Mercedes fleet\n'
    '★ Duration : 8 hours (with driver) or 24 hours (vehicle-only)\n'
    '★ Program\n'
    '   ▶ Options\n'
    '   • Vehicle + driver\n'
    '   • Vehicle-only 24h — 55,000 KRW own insurance per day\n'
    '   ▶ Add-ons\n'
    '   • English-speaking driver (additional surcharge)\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Driver meals/accommodation (billed after event)'
),
'#P-561': (
    '★ VIP Charter Service\n'
    '★ Vehicle : Premium Mercedes fleet\n'
    '★ Duration : 8 hours (with driver) or 24 hours (vehicle-only)\n'
    '★ Program\n'
    '   ▶ Options\n'
    '   • Vehicle + driver\n'
    '   • Vehicle-only 24h — 55,000 KRW own insurance per day\n'
    '   ▶ Add-ons\n'
    '   • English-speaking driver (additional surcharge)\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Driver meals/accommodation (billed after event)'
),
'#P-562': (
    '★ VIP Charter Service\n'
    '★ Vehicle : Premium Mercedes fleet\n'
    '★ Duration : 8 hours (with driver) or 24 hours (vehicle-only)\n'
    '★ Program\n'
    '   ▶ Options\n'
    '   • Vehicle + driver\n'
    '   • Vehicle-only 24h — 55,000 KRW own insurance per day\n'
    '   ▶ Add-ons\n'
    '   • English-speaking driver (additional surcharge)\n'
    '   ▶ Not Included\n'
    '   • Fuel • Tolls • Driver meals/accommodation (billed after event)'
),
}

wb = openpyxl.load_workbook('E:/Interview Co/10. Agent SaaS/Code/agent-saas/data/products_master_v24.xlsx')

updated = 0
for shname in wb.sheetnames:
    ws = wb[shname]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        pn_col = headers.index('product_number') + 1
        desc_col = headers.index('description') + 1
    except ValueError:
        continue
    for r in range(2, ws.max_row + 1):
        pn = ws.cell(r, pn_col).value
        if pn in REWRITES:
            ws.cell(r, desc_col).value = REWRITES[pn]
            updated += 1

wb.save('E:/Interview Co/10. Agent SaaS/Code/agent-saas/data/products_master_v24.xlsx')
print(f'Updated {updated} rows -> v24 saved with {len(REWRITES)} unique product_numbers reformatted.')
