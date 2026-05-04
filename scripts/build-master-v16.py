"""Build products_master_v16.xlsx — clean up v15 data quality issues.

Fixes:
  - Dr.Tune's: prices were stored in USD but currency=USD got lost. Convert
    to KRW @ 1500 KRW/USD reference rate.
  - Nest Clinic 5 rows that came from v11 with name=None (Excel merged
    cells leaked through): assign proper base names from their description,
    and group as variants:
      * "Dermatological procedures" — 3 tiers (4M / 6M / 8M)
      * "Full Body Package" — 3 protocols (5M / 8M / 15M)
"""
from openpyxl import load_workbook

USD_TO_KRW = 1500  # rough; same rate used elsewhere in the app

wb = load_workbook('data/products_master_v15.xlsx')
ws = wb['Products']
header = [c.value for c in ws[1]]
ic = lambda n: header.index(n)

# ── Fix Dr.Tune's USD → KRW ──────────────────────────────────────────────
fixed_dr = 0
for row in range(2, ws.max_row + 1):
    partner = ws.cell(row=row, column=ic('partner_short') + 1).value
    if partner == "Dr.Tune's":
        cur = ws.cell(row=row, column=ic('price_currency') + 1).value
        price = ws.cell(row=row, column=ic('base_price') + 1).value
        if price and price < 100000:  # USD-magnitude (under $100K) — assume USD
            ws.cell(row=row, column=ic('base_price') + 1, value=int(round(price * USD_TO_KRW)))
            ws.cell(row=row, column=ic('price_currency') + 1, value='KRW')
            # Clear bogus variant_label that was carried over from v15's regex
            # (they had no real variant — Glow/Premium Anti-Aging etc are
            # distinct products, not variants).
            ws.cell(row=row, column=ic('variant_label') + 1, value=None)
            fixed_dr += 1
print(f'Dr.Tune\'s prices converted: {fixed_dr}')

# ── Fix Nest 5 misnamed rows ─────────────────────────────────────────────
# Map: (price → correct name, variant_label, grade)
NEST_PATCHES = {
    # Dermatological procedures — 3 tiers
    4_000_000:  ('Dermatological Procedures', 'Tier 1 — Ulthera 300 + Thermage 600 + SkinCeuticals', 'Standard'),
    6_000_000:  ('Dermatological Procedures', 'Tier 2 — Thermage 600 + Eye Thermage 450 + Rejuran', 'Premium'),
    8_000_000:  ('Dermatological Procedures', 'Tier 3 — Ulthera 600 + Thermage 600 + Rejuran HB',  'VIP Premium'),
    # Full Body Package — 3 protocols
    5_000_000:  ('Full Body Package', 'Antiaging + Collagen Volumizing (Thermage 600 + body)',     'Standard'),
    # Note: 8M for Nest occurs twice (one Derm, one Body). Distinguish by description.
    15_000_000: ('Full Body Package', 'Slimming + Total Antiaging (Body Thermage 3000, ≤5 sites)', 'VIP Premium'),
}

fixed_nest = 0
for row in range(2, ws.max_row + 1):
    partner = ws.cell(row=row, column=ic('partner_short') + 1).value
    if partner != 'Nest Clinic':
        continue
    name = ws.cell(row=row, column=ic('name') + 1).value
    desc = ws.cell(row=row, column=ic('description') + 1).value or ''
    price = ws.cell(row=row, column=ic('base_price') + 1).value

    # The 8M Nest row is ambiguous (Dermatological vs Body) — use description to disambiguate
    if price == 8_000_000:
        if 'Body' in desc or 'Slimming' in desc:
            ws.cell(row=row, column=ic('name') + 1, value='Full Body Package')
            ws.cell(row=row, column=ic('variant_label') + 1, value='Slimming + Antiaging (Body Thermage 1200, ≤2 sites)')
            ws.cell(row=row, column=ic('grade') + 1, value='Premium')
            fixed_nest += 1
        elif 'Ulthera' in desc or 'Thermage' in desc:
            # already handled by NEST_PATCHES[8M]
            new_name, var_label, grade = NEST_PATCHES[8_000_000]
            ws.cell(row=row, column=ic('name') + 1, value=new_name)
            ws.cell(row=row, column=ic('variant_label') + 1, value=var_label)
            ws.cell(row=row, column=ic('grade') + 1, value=grade)
            fixed_nest += 1
        continue

    # Other prices: only patch if name doesn't already match what we want
    if price in NEST_PATCHES:
        new_name, var_label, grade = NEST_PATCHES[price]
        # If name is one of the misnamed values (DIAR Secret Injection, Juvederm, Domestic Filler)
        # OR matches an inherited name from another partner, replace.
        if name in (None, '', 'DIAR Secret Injection', 'Juvederm', 'Domestic Filler') or name == new_name:
            ws.cell(row=row, column=ic('name') + 1, value=new_name)
            ws.cell(row=row, column=ic('variant_label') + 1, value=var_label)
            ws.cell(row=row, column=ic('grade') + 1, value=grade)
            fixed_nest += 1

print(f'Nest rows patched: {fixed_nest}')

# Also tag the Nest "Dermatological procedures" first-tier row that had grade='Dermatological procedures'
for row in range(2, ws.max_row + 1):
    partner = ws.cell(row=row, column=ic('partner_short') + 1).value
    if partner != 'Nest Clinic':
        continue
    name = ws.cell(row=row, column=ic('name') + 1).value
    grade = ws.cell(row=row, column=ic('grade') + 1).value
    if name == 'Dermatological procedures' and grade == 'Dermatological procedures':
        ws.cell(row=row, column=ic('name') + 1, value='Dermatological Procedures')
        ws.cell(row=row, column=ic('grade') + 1, value='Standard')
        # variant_label stays as set above (Tier 1)
    if name == 'Full Body Package' and grade == 'Full Body Package':
        ws.cell(row=row, column=ic('grade') + 1, value='Standard')

wb.save('data/products_master_v16.xlsx')
print('Saved data/products_master_v16.xlsx')

# Verify
wb2 = load_workbook('data/products_master_v16.xlsx', data_only=True)
ws2 = wb2['Products']
print('\nNest Clinic rows in v16:')
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[ic('partner_short')] == 'Nest Clinic':
        print(f"  name='{r[ic('name')]}' | variant='{r[ic('variant_label')]}' | grade='{r[ic('grade')]}' | price={r[ic('base_price')]}")
print("\nDr.Tune's rows in v16:")
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[ic('partner_short')] == "Dr.Tune's":
        print(f"  name='{r[ic('name')]}' | variant='{r[ic('variant_label')]}' | price={r[ic('base_price')]} {r[ic('price_currency')]}")
