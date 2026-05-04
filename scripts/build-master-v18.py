"""Build products_master_v18.xlsx from v17.

Two consolidations:

  1. DIAR (159 rows → 11 base products + 159 variants).
     Product number ranges drive the bucket assignment:
       P-021..023  DIAR Membership          (3)
       P-024..032  MCT Stem Cell Therapy    (9)
       P-033..054  Tension Lifting          (22)
       P-055..071  Skin Booster             (17)
       P-072..074  Filler                   (3)
       P-075..113  Botox                    (39)
       P-114..120  Body Filler              (7)
       P-121..125  DIAR Signature Care      (5)
       P-126..133  DIAR Private Programs    (8)
       P-134..149  Acne / Pigment / Pore / Peel (16)
       P-150..179  Hair Removal — Face & Body (30)

     Each row's existing `name` (treatment name) and `variant_label`
     (spec/brand/sessions) get composed into the new `variant_label`
     under the bucket's product name.

  2. Hotel (37 rows → 13 base products + 37 variants).
     Group by partner_short, product name = partner_short. Each row's
     room name (cleaned) becomes the variant_label.

     Cleanup rules applied to room names:
       - strip " [★5 Hotel]" suffix (every row had it = no info)
       - " [Hanok]" → " (Hanok)" (information-bearing, keep inline)
       - strip grade words in parens: (Deluxe) (Suite) (Special Suite)
         (Standard) (Premium) (VIP Premium) (VVIP) (Hanok-when-redundant)
       - leave descriptive parens alone: (Lake View), (Room + Dining),
         (Minagi Omakase Set), etc.

  All other rows pass through unchanged.
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook, Workbook

SRC = 'data/products_master_v17.xlsx'
DST = 'data/products_master_v18.xlsx'

# ── Bucket mapping for DIAR ──────────────────────────────────────────────
DIAR_RANGES = [
    (21, 23,  'DIAR Membership'),
    (24, 32,  'MCT Stem Cell Therapy'),
    (33, 54,  'Tension Lifting'),
    (55, 71,  'Skin Booster'),
    (72, 74,  'Filler'),
    (75, 113, 'Botox'),
    (114, 120,'Body Filler'),
    (121, 125,'DIAR Signature Care'),
    (126, 133,'DIAR Private Programs'),
    (134, 149,'Acne / Pigment / Pore / Peel'),
    (150, 179,'Hair Removal — Face & Body'),
]

def diar_bucket_for(num_int):
    for lo, hi, name in DIAR_RANGES:
        if lo <= num_int <= hi:
            return name
    return None

def diar_variant_label(bucket, name, var):
    name = (name or '').strip()
    var = (var or '').strip()
    if bucket == 'DIAR Membership':
        # name="DIAR Membership", var="Red"/"Purple"/"Black"
        return var or name
    if bucket == 'MCT Stem Cell Therapy':
        # name="MCT Stem Cell — Face" / "...Face + Stem-Cell Fusion" / "...Scalp"
        sub = re.sub(r'^MCT Stem Cell\s*[—–-]\s*', '', name)
        return f'{sub} · {var}' if var else sub
    # default: combine treatment name with spec/brand/sessions
    if var:
        return f'{name} · {var}'
    return name

# ── Hotel cleanup ────────────────────────────────────────────────────────
GRADE_PARENS = re.compile(
    r'\s*\((?:Deluxe|Suite|Special Suite|Standard|Premium|VIP Premium|VVIP)\)\s*'
)

def hotel_variant_label(name):
    s = (name or '').strip()
    # bracket suffix cleanup. Hanok is information-bearing — convert to inline
    # paren AFTER grade-paren stripping so it doesn't get caught.
    s = re.sub(r'\s*\[★5\s*Hotel\]\s*$', '', s)
    has_hanok = bool(re.search(r'\s*\[Hanok\]\s*$', s))
    s = re.sub(r'\s*\[Hanok\]\s*$', '', s)
    # strip grade-word parens (Deluxe/Suite/Special Suite/etc — not Hanok)
    s = GRADE_PARENS.sub(' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    if has_hanok:
        s = f'{s} (Hanok)'
    return s

# Partner short → canonical product name (used as `name` column for the consolidated
# hotel product). Keep partner_short as-is for now; the formal partner_name carries
# the long official spelling.
HOTEL_PRODUCT_NAME = {
    'Shilla':            'Shilla Seoul',
    'Signiel':           'Signiel Seoul',
    'Four Seasons':      'Four Seasons Seoul',
    'Park Hyatt Seoul':  'Park Hyatt Seoul',
    'InterContinental':  'InterContinental Seoul',
    'Paradise':          'Paradise City',
    'INSPIRE':           'Inspire Entertainment Resort',
    'Park Hyatt Incheon':'Park Hyatt Incheon',
    'Oakwood Incheon':   'Oakwood Premier Incheon',
    'Sheraton':          'Sheraton',
    'Hanok Essay':       'Hanok Essay Hahoe',
    'Gyeongwonjae':      'Gyeongwonjae Ambassador',
    'Sofitel':           'Sofitel Ambassador Seoul',
}

# ── Main ─────────────────────────────────────────────────────────────────
def main():
    wb = load_workbook(SRC)
    ws = wb['Products']
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    diar_rows = []   # collected, will be re-emitted
    hotel_rows_by_partner = {}  # partner_short -> [rows]

    for row in ws.iter_rows(min_row=2, values_only=True):
        partner_short = row[idx['partner_short']]
        subcat = row[idx['subcategory']]
        if partner_short == 'DIAR':
            diar_rows.append(list(row))
        elif subcat == 'Hotel':
            hotel_rows_by_partner.setdefault(partner_short, []).append(list(row))
        else:
            out_rows.append(list(row))

    # ── DIAR consolidation ──────────────────────────────────────────────
    # Group by bucket, preserving order.
    bucket_rows = {}  # bucket_name -> [(orig_row, new_variant_label)]
    bucket_order = []
    for r in diar_rows:
        pn = r[idx['product_number']]
        m = re.match(r'#P-0*(\d+)', pn or '')
        if not m:
            print(f'WARN: bad product_number {pn!r}', file=sys.stderr)
            continue
        num = int(m.group(1))
        bucket = diar_bucket_for(num)
        if not bucket:
            print(f'WARN: no bucket for {pn} ({r[idx["name"]]})', file=sys.stderr)
            continue
        new_var = diar_variant_label(bucket, r[idx['name']], r[idx['variant_label']])
        if bucket not in bucket_rows:
            bucket_rows[bucket] = []
            bucket_order.append(bucket)
        bucket_rows[bucket].append((r, new_var))

    # Emit DIAR rows. Update name + variant_label + subcategory.
    # subcategory was 'DIAR'; keep as 'DIAR' so subcategory dropdown stays.
    for bucket in bucket_order:
        for orig, new_var in bucket_rows[bucket]:
            new_row = list(orig)
            new_row[idx['name']] = bucket
            new_row[idx['variant_label']] = new_var
            new_row[idx['grade']] = None  # selection grade is meaningless after consolidation
            out_rows.append(new_row)

    # ── Hotel consolidation ─────────────────────────────────────────────
    # Order: per HOTEL_PRODUCT_NAME insertion, then by row sort_order.
    # Disambiguate: if multiple input rows under one partner produce the same
    # variant_label (e.g. INSPIRE has "Sun Tower Inspire Suite" twice with
    # different prices), suffix " (#2)", " (#3)" so the upload UPSERT keeps
    # both as distinct variants.
    for partner_short in HOTEL_PRODUCT_NAME:
        rows = hotel_rows_by_partner.get(partner_short, [])
        if not rows:
            continue
        product_name = HOTEL_PRODUCT_NAME[partner_short]
        seen_labels = {}
        for r in rows:
            new_row = list(r)
            label = hotel_variant_label(r[idx['name']])
            seen_labels[label] = seen_labels.get(label, 0) + 1
            if seen_labels[label] > 1:
                label = f'{label} (#{seen_labels[label]})'
            new_row[idx['name']] = product_name
            new_row[idx['variant_label']] = label
            new_row[idx['grade']] = None
            out_rows.append(new_row)
    # Catch any partner not in our explicit list (defensive)
    for partner_short, rows in hotel_rows_by_partner.items():
        if partner_short in HOTEL_PRODUCT_NAME:
            continue
        product_name = partner_short or 'Unknown Hotel'
        print(f'WARN: hotel partner {partner_short!r} not in HOTEL_PRODUCT_NAME, using fallback', file=sys.stderr)
        for r in rows:
            new_row = list(r)
            new_row[idx['name']] = product_name
            new_row[idx['variant_label']] = hotel_variant_label(r[idx['name']])
            new_row[idx['grade']] = None
            out_rows.append(new_row)

    # ── Write v18 ───────────────────────────────────────────────────────
    out = Workbook()
    ws_out = out.active
    ws_out.title = 'Products'
    ws_out.append(hdr)
    for r in out_rows:
        ws_out.append(r)

    # Summary sheet
    ws_sum = out.create_sheet('Summary')
    ws_sum.append(['Section', 'Bucket / Partner', 'Variant count'])
    for bucket in bucket_order:
        ws_sum.append(['DIAR', bucket, len(bucket_rows[bucket])])
    for partner_short, rows in hotel_rows_by_partner.items():
        ws_sum.append(['Hotel', HOTEL_PRODUCT_NAME.get(partner_short, partner_short), len(rows)])

    out.save(DST)
    print(f'Wrote {DST}')
    print(f'  Total rows: {len(out_rows)}')
    print(f'  DIAR: {sum(len(v) for v in bucket_rows.values())} variants in {len(bucket_order)} products')
    print(f'  Hotel: {sum(len(v) for v in hotel_rows_by_partner.values())} variants in {len(hotel_rows_by_partner)} products')

if __name__ == '__main__':
    main()
