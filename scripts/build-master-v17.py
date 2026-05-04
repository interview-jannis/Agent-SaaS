"""Build products_master_v17.xlsx from v16.

Three changes:
  1. Strip selection-grade parens from name: "(Silver)" / "(Gold)" / "(Diamond)"
     (these are Selection Sheet grades, not part of the actual product name).
     Other parenthetical qualifiers — (Suite), (Diamond Premium), (Face),
     (full-face), etc — are kept verbatim.
  2. Merge gender into variants: rows whose name ends with " · Male" or
     " · Female" get the suffix stripped and `variant_label` set to "Male"
     or "Female". The upload API groups (category, partner, name) into a
     product family and matches variants by variant_label, so this turns
     20 K-Medical rows into 10 base products + 20 variants.
  3. Restore description newlines for K-Medical by re-pulling from the
     Internal_Price source sheet (which preserves '*' / '•' bulleted lines).
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

SRC = 'data/products_master_v16.xlsx'
DST = 'data/products_master_v17.xlsx'
INTERNAL = 'data/Internal_Price sheet for Proposal_draft_2_v7_26.02.24.xlsx'

# ── 1. Build K-Medical description map from source ────────────────────────
# Source sheet `1. K-Medical` rows where col C (Grade) is set carry the
# Check-up Detail (col H) for that (partner, grade). Partner is in col B
# but is set on the FIRST row of each partner's block — subsequent grade
# rows under the same partner have B=None. We forward-fill it.
def load_kmedical_source():
    wb = load_workbook(INTERNAL, data_only=True)
    ws = wb['1. K-Medical']
    by_partner_grade = {}
    current_partner_full = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: A Package, B Partners, C Grade, D Selection Grade,
        # E Quotation(KRW), F USD, G Time, H Check-up Detail, I Info
        partner_cell = row[1]
        grade = row[2]
        detail = row[7]
        if partner_cell:
            current_partner_full = partner_cell
        if grade and detail and current_partner_full:
            # Use the first non-empty line of partner cell (e.g. "Gil Hospital")
            partner_first = str(current_partner_full).split('\n')[0].strip()
            by_partner_grade[(partner_first, str(grade).strip())] = str(detail)
    return by_partner_grade

# Map v16 partner_short → source partner first-line
PARTNER_SHORT_TO_SOURCE = {
    'Gil Hospital':      'Gil Hospital',
    'SNUH Gangnam':      'SNUH healthcare System',
    'Catholic Hospital': 'The Catholic University of Korea',
    'Asan Medical':      'Asan Medical Center',
}

src_desc = load_kmedical_source()
print(f'K-Medical source descriptions loaded: {len(src_desc)}')

# ── 2. Process v16 → v17 ──────────────────────────────────────────────────
wb = load_workbook(SRC)
ws = wb['Products']
header = [c.value for c in ws[1]]
ic = {h: i for i, h in enumerate(header)}

SELECTION_GRADE_PARENS = re.compile(r'\s*\((?:Silver|Gold|Diamond)\)\s*')
GENDER_SUFFIX = re.compile(r'\s*[·•]\s*(Male|Female)\s*$')

stats = {'name_paren_stripped': 0, 'gender_extracted': 0, 'desc_restored': 0}

for row_idx in range(2, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=ic['name'] + 1).value
    if not name:
        continue

    # 2a — extract gender suffix and set variant_label
    m = GENDER_SUFFIX.search(name)
    if m:
        gender = m.group(1)
        name = GENDER_SUFFIX.sub('', name)
        # Only overwrite variant_label if it's empty — preserve any existing
        # tier label (no current rows have both, but defensive).
        existing_var = ws.cell(row=row_idx, column=ic['variant_label'] + 1).value
        if not existing_var:
            ws.cell(row=row_idx, column=ic['variant_label'] + 1, value=gender)
        stats['gender_extracted'] += 1

    # 2b — strip selection grade parens
    new_name = SELECTION_GRADE_PARENS.sub(' ', name).strip()
    new_name = re.sub(r'\s+', ' ', new_name)
    if new_name != name:
        stats['name_paren_stripped'] += 1
    ws.cell(row=row_idx, column=ic['name'] + 1, value=new_name)

    # 2c — restore K-Medical description with \n from source
    cat = ws.cell(row=row_idx, column=ic['category'] + 1).value
    if cat == 'K-Medical':
        partner_short = ws.cell(row=row_idx, column=ic['partner_short'] + 1).value
        grade = ws.cell(row=row_idx, column=ic['grade'] + 1).value
        src_partner = PARTNER_SHORT_TO_SOURCE.get(partner_short)
        if src_partner and grade:
            key = (src_partner, str(grade).strip())
            if key in src_desc:
                ws.cell(row=row_idx, column=ic['description'] + 1, value=src_desc[key])
                stats['desc_restored'] += 1

print('Stats:', stats)

wb.save(DST)
print(f'Saved {DST}')

# ── Verify ────────────────────────────────────────────────────────────────
wb2 = load_workbook(DST, data_only=True)
ws2 = wb2['Products']
print('\n=== K-Medical rows in v17 ===')
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[ic['category']] == 'K-Medical':
        d = r[ic['description']] or ''
        print(f"{r[ic['partner_short']]:20} | {r[ic['name']]:30} | var={r[ic['variant_label']]} | nl={'YES' if chr(10) in d else 'NO'} | {d[:80].replace(chr(10),' / ')}")
