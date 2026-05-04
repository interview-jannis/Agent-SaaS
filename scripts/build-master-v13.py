"""Build products_master_v13.xlsx from v12 + add rows from folder data sources.

Adds:
- Ruby Clinic (K-Beauty) — 4 packages from mail
- Dekabi Stem Cell Clinic (K-Beauty) — 1 placeholder (price TBD)
- SOFITEL Ambassador Seoul (Subpackage Hotel) — 5 rooms + 3 packages (TBD)
- World K-POP Center "Special" (K-Starcation) — 4th custom-quote tier

Skipped:
- DIAR (PDF is image-based; OCR or manual entry needed later)
"""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook('data/products_master_v12.xlsx')
ws = wb['Products']
header = [c.value for c in ws[1]]
ic = lambda n: header.index(n)

last_num = 0
cat_max_sort = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    pn = r[ic('product_number')]
    if pn and isinstance(pn, str) and pn.startswith('#P-'):
        try:
            last_num = max(last_num, int(pn[3:]))
        except ValueError:
            pass
    cat = r[ic('category')]
    so = r[ic('sort_order')]
    if cat and isinstance(so, (int, float)):
        cat_max_sort[cat] = max(cat_max_sort.get(cat, 0), int(so))

print(f"Last #P-: {last_num}")
print(f"Max sort per cat: {cat_max_sort}")

def empty_row():
    return {h: '' for h in header}

new_rows = []

# Ruby Clinic
ruby_base = {
    'category': 'K-Beauty',
    'subcategory': 'K-Beauty Program',
    'partner_name': 'Ruby Clinic',
    'partner_short': 'Ruby',
    'price_currency': 'KRW',
    'duration_value': 1, 'duration_unit': 'days',
    'why_recommendation': 'Premium dermatology clinic specializing in customized VIP packages.',
    'location_address': 'Seoul',
    'contact_email': 'jangjaehee17@gmail.com',
    'has_prayer_room': '', 'dietary_type': 'none',
    'is_active': 'TRUE',
    'source_sheet': 'Ruby/Ruby_mail.txt (2026-04-23)',
}
ruby_pkgs = [
    ('Custermize Lifting Package', 'Premium', 22_000_000,
     'Ultherapy Prime + Thermage + ONDA Lifting + Full face Botox + Hyperbaric oxygen therapy + Face exfoliation + LDM + Recovery therapy + Customized IV drip + Modeling mask pack'),
    ('Custermize Pigmentation Package', 'Premium', 18_000_000,
     '5 types of pigmentation laser + Regeneration & Whitening laser + Full face skin booster + Full face ECM booster + Hyperbaric oxygen therapy + Face exfoliation + LDM + Recovery therapy + Customized IV drip + Modeling mask pack'),
    ('Oxygen Care Package', 'Premium', 16_000_000,
     'Astronaut therapy (full body care) + Oxygen exfoliation + Hyperbaric oxygen therapy + Hair therapy (shampoo, exosome, growth factor injection, oxygen therapy) + LDM + Recovery therapy + Customized IV drip + Modeling mask pack'),
    ('Autologous Blood Stem Cell Therapy Package', 'Premium', 15_000_000,
     'Face exfoliation + IV + skin + scalp (240cc) + LDM + Recovery therapy + Hyperbaric oxygen therapy'),
]
sort_order = cat_max_sort.get('K-Beauty', 0)
for name, grade, price, desc in ruby_pkgs:
    last_num += 1
    sort_order += 1
    r = empty_row()
    r.update(ruby_base)
    r['product_number'] = f'#P-{last_num:03d}'
    r['name'] = name
    r['grade'] = grade
    r['base_price'] = price
    r['description'] = desc
    r['sort_order'] = sort_order
    new_rows.append(r)

# Dekabi placeholder
last_num += 1
sort_order += 1
new_rows.append({**empty_row(), **{
    'product_number': f'#P-{last_num:03d}',
    'category': 'K-Beauty',
    'subcategory': 'K-Beauty Program',
    'partner_name': 'Dekabi Stem Cell Clinic',
    'partner_short': 'Dekabi',
    'name': 'Dekabi Stem Cell Regenerative Care (TBD)',
    'grade': 'Premium',
    'price_currency': 'KRW',
    'duration_value': 1, 'duration_unit': 'days',
    'description': 'One-day stem cell procedure for chronic conditions; minimum 1-month effect window per session; recommended for patients seeking shorter recovery vs traditional 3+ month treatment cycles.',
    'why_recommendation': 'Regenerative stem cell clinic; reportedly 4-month, 2-stage protocol; sought after by global VIPs.',
    'location_address': 'Seoul',
    'has_prayer_room': '', 'dietary_type': 'none',
    'is_active': 'FALSE',
    'sort_order': sort_order,
    'source_sheet': 'Dekabi/Dekabi_Mail.txt',
    'notes': 'Pricing/specific package details pending — contact partner.',
}})

# SOFITEL rooms
sof_base = {
    'category': 'Subpackage',
    'subcategory': 'Hotel',
    'partner_name': 'SOFITEL Ambassador Seoul',
    'partner_short': 'Sofitel',
    'price_currency': 'KRW',
    'duration_value': 1, 'duration_unit': 'days',
    'why_recommendation': 'French luxury hotel brand with Muslim-friendly amenities (prayer mat, qibla indicator, Arabic coffee, dedicated Guest Relations) — well-suited for GCC VIP clients.',
    'location_address': 'Songpa-gu, Seoul (Lake View)',
    'has_prayer_room': 'TRUE', 'dietary_type': 'halal_friendly',
    'is_active': 'TRUE',
    'source_sheet': 'Hotel/SOFITEL/(주)인터뷰_중동VIP_객실패키지_구성 260317_1st Draft.pdf',
    'notes': 'Subject to 10% VAT. base_price = Special Rate (negotiated); price_max = Published Rate.',
    'price_unit': 'per night',
}
sof_max = cat_max_sort.get('Subpackage', 0)
sof_rooms = [
    ('Luxury Lake Room (Lake View)', 'Standard', 360_000, 1_100_000, '34 sqm, King or Twin bed.'),
    ('Prestige Suite (Lake View)', 'Premium', 700_000, 2_000_000, '70 sqm, King bed; Executive Lounge access for 2 included.'),
    ('Opera Suite (Panoramic Lake View)', 'VIP Premium', 1_500_000, 5_000_000, '135 sqm, Super King bed; Executive Lounge access for 2 included.'),
    ('Presidential Suite (Panoramic Lake View)', 'VVIP', '', 22_000_000, '244 sqm, two Super King beds; Executive Lounge access for 2 included. Special Rate by Sales Manager inquiry.'),
    ('1-Bedroom Premier Suite (Lake View)', 'Premium', 500_000, 2_000_000, '71 sqm Residence, King bed.'),
]
for name, grade, sp, pub, desc in sof_rooms:
    last_num += 1
    sof_max += 1
    r = empty_row()
    r.update(sof_base)
    r['product_number'] = f'#P-{last_num:03d}'
    r['name'] = name
    r['grade'] = grade
    r['base_price'] = sp if sp != '' else pub
    r['price_min'] = sp if sp != '' else ''
    r['price_max'] = pub
    r['description'] = desc
    r['sort_order'] = sof_max
    new_rows.append(r)

# SOFITEL packages (3) — TBD pricing
sof_pkgs = [
    ('ROYAL GOLD Package (Room + Dining)', 'VIP Premium',
     'Noor Royal Indulgence Collection · Royal Arabian Indulgence — guest room + in-room dining (3-course menu). All packages include Yoo Jian-jak handcrafted najeon (mother-of-pearl) gift box.'),
    ('ROYAL Platinum Package (Room + Wellness)', 'VIP Premium',
     'Serenity Sanctuary Escape · Safa Serenity Retreat — guest room + Premium Spa & Wellness experience (Biologique Recherche, Thermes Marins by certified therapists). All packages include Yoo Jian-jak handcrafted najeon gift box.'),
    ('ROYAL Beauty Package (Room + K-Beauty)', 'VIP Premium',
     'K-Beauty Prestige Escape — guest room + exclusive K-Beauty experience (premium devices, customized consulting, gift set). All packages include Yoo Jian-jak handcrafted najeon gift box.'),
]
for name, grade, desc in sof_pkgs:
    last_num += 1
    sof_max += 1
    r = empty_row()
    r.update(sof_base)
    r['product_number'] = f'#P-{last_num:03d}'
    r['name'] = name
    r['grade'] = grade
    r['description'] = desc
    r['sort_order'] = sof_max
    r['is_active'] = 'FALSE'
    r['notes'] = 'Package pricing TBD - contact Sofitel sales for quote.'
    r['base_price'] = ''
    r['price_min'] = ''
    r['price_max'] = ''
    new_rows.append(r)

# World K-POP Special
last_num += 1
ks_max = cat_max_sort.get('K-Starcation', 0) + 1
new_rows.append({**empty_row(), **{
    'product_number': f'#P-{last_num:03d}',
    'category': 'K-Starcation',
    'subcategory': 'K-POP Camp',
    'partner_name': 'World K-POP Center',
    'partner_short': 'World K-POP',
    'name': 'K-POP CAMP - Special (Custom)',
    'grade': 'VVIP',
    'price_currency': 'KRW',
    'description': 'Customized K-pop experience tailored to VVIP guests - premier idol meet-up, music video / styling shoot, custom curation. Quoted per request.',
    'why_recommendation': 'For guests who want a fully bespoke K-pop experience beyond the Basic/Gold/Platinum tiers.',
    'location_address': 'Seoul',
    'contact_email': 'w-kpop@naver.com',
    'has_prayer_room': '', 'dietary_type': 'none',
    'is_active': 'FALSE',
    'sort_order': ks_max,
    'source_sheet': 'K-Starcation/K-STARCATION program xlsx',
    'notes': 'Custom quote - pricing per scope.',
}})

# Append
start_row = ws.max_row + 1
for i, row in enumerate(new_rows):
    for col_idx, h in enumerate(header, 1):
        ws.cell(row=start_row + i, column=col_idx, value=row.get(h, ''))

# Recompute Summary
counts = Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    v = r[ic('category')]
    if v: counts[v] += 1
summary = wb['Summary']
for row in summary.iter_rows(min_row=2):
    for c in row:
        c.value = None
ORDER = ['K-Medical', 'K-Beauty', 'K-Wellness', 'K-Education', 'K-Starcation', 'Subpackage']
total = 0
ri = 2
for cat in ORDER:
    n = counts.get(cat, 0)
    summary.cell(row=ri, column=1, value=cat)
    summary.cell(row=ri, column=2, value=n)
    total += n
    ri += 1
summary.cell(row=ri, column=1, value='TOTAL')
summary.cell(row=ri, column=2, value=total)

wb.save('data/products_master_v13.xlsx')
print(f"\nAdded {len(new_rows)} rows. v13 totals:")
for cat in ORDER:
    print(f"  {cat}: {counts.get(cat, 0)}")
print(f"  TOTAL: {total}")
