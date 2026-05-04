"""Import products_master_v16.xlsx into Supabase.

Workflow:
  1. Wipe product_variants and products (clean import).
  2. Drop legacy duplicate categories ("Medical"/"Beauty"/"Wellness" without K-).
  3. Ensure K-prefixed categories + Subpackage exist.
  4. Insert product_subcategories from unique (cat, sub) pairs in v16.
  5. Insert one product per (cat, sub, partner, name) family.
  6. Insert one variant per row, with rule: solo families get NULL label.

Safe to re-run. For dev/test only.
"""
import os
import sys
import json
import urllib.request
import urllib.error
from urllib.parse import urlencode
from collections import defaultdict
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# ── Env ──────────────────────────────────────────────────────────────────
env = {}
with open('.env.local') as f:
    for line in f:
        if '=' in line and not line.strip().startswith('#'):
            k, v = line.strip().split('=', 1)
            env[k] = v
URL = env['NEXT_PUBLIC_SUPABASE_URL']
KEY = env['SUPABASE_SERVICE_ROLE_KEY']

def supa(method, path, body=None, params=None):
    url = f'{URL}/rest/v1/{path}'
    if params:
        url += '?' + urlencode(params)
    req = urllib.request.Request(url, method=method)
    req.add_header('apikey', KEY)
    req.add_header('Authorization', f'Bearer {KEY}')
    req.add_header('Content-Type', 'application/json')
    req.add_header('Prefer', 'return=representation')
    data = json.dumps(body, ensure_ascii=False).encode() if body is not None else None
    try:
        with urllib.request.urlopen(req, data=data) as r:
            txt = r.read().decode()
            return json.loads(txt) if txt else None
    except urllib.error.HTTPError as e:
        msg = e.read().decode()
        print(f'HTTP {e.code} on {method} {path}: {msg[:600]}')
        if body:
            print(f'  body: {json.dumps(body, ensure_ascii=False)[:400]}')
        raise

# ── 1. Wipe ──────────────────────────────────────────────────────────────
print('Wiping product_variants...')
supa('DELETE', 'product_variants', params={'id': 'gt.00000000-0000-0000-0000-000000000000'})
print('Wiping products...')
supa('DELETE', 'products', params={'id': 'gt.00000000-0000-0000-0000-000000000000'})
print('Wiping product_subcategories...')
supa('DELETE', 'product_subcategories', params={'id': 'gt.00000000-0000-0000-0000-000000000000'})

# ── 2. Drop legacy non-K categories (if no products reference them) ─────
print('Dropping legacy categories...')
existing = supa('GET', 'product_categories', params={'select': 'id,name'})
for c in existing or []:
    if c['name'] in ('Medical', 'Beauty', 'Wellness'):
        try:
            supa('DELETE', 'product_categories', params={'id': f'eq.{c["id"]}'})
            print(f'  removed legacy "{c["name"]}"')
        except Exception:
            print(f'  could not remove "{c["name"]}" (may have referencing rows)')

# ── 3. Ensure K-prefixed + Subpackage categories with right sort_order ──
CATS = ['K-Medical', 'K-Beauty', 'K-Wellness', 'K-Education', 'K-Starcation', 'Subpackage']
existing = supa('GET', 'product_categories', params={'select': 'id,name'})
cat_map = {c['name']: c['id'] for c in (existing or [])}
for i, name in enumerate(CATS):
    if name not in cat_map:
        result = supa('POST', 'product_categories', body={'name': name, 'sort_order': i})
        cat_map[name] = result[0]['id']
        print(f'  + created {name}')
    else:
        supa('PATCH', 'product_categories', body={'sort_order': i}, params={'id': f'eq.{cat_map[name]}'})

# ── 4. Read v16 and gather unique subcategories per category ────────────
wb = load_workbook('data/products_master_v16.xlsx', data_only=True)
ws = wb['Products']
header = [c.value for c in ws[1]]
ic = lambda n: header.index(n)
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[ic('product_number')]]
print(f'Read {len(rows)} rows from v16')

# Unique (cat, sub) pairs
sub_pairs = sorted({(r[ic('category')], r[ic('subcategory')] or '') for r in rows
                    if r[ic('category')] in cat_map and r[ic('subcategory')]})

print(f'Inserting {len(sub_pairs)} subcategories...')
sub_map = {}  # (cat, sub) -> id
for idx, (cat, sub) in enumerate(sub_pairs):
    body = {'name': sub, 'category_id': cat_map[cat], 'sort_order': idx}
    result = supa('POST', 'product_subcategories', body=body)
    sub_map[(cat, sub)] = result[0]['id']

# ── 5. Group rows by (cat, sub, partner_short, name) ────────────────────
groups = defaultdict(list)
for r in rows:
    key = (r[ic('category')],
           r[ic('subcategory')] or '',
           r[ic('partner_short')] or r[ic('partner_name')] or '',
           r[ic('name')] or '')
    groups[key].append(r)
print(f'Grouped into {len(groups)} product families')

def normalize_bool(v):
    if v is True or v in ('TRUE', 'true', True): return True
    if v is False or v in ('FALSE', 'false', False): return False
    return None

def normalize_dietary(v):
    valid = {'halal_certified', 'halal_friendly', 'muslim_friendly', 'pork_free', 'none'}
    return v if v in valid else 'none'

# Sort families deterministically for product_number assignment
def gkey(item):
    (cat, sub, partner, name), grp = item
    try:
        ci = CATS.index(cat)
    except ValueError:
        ci = len(CATS)
    so = grp[0][ic('sort_order')]
    so = so if isinstance(so, (int, float)) else 999999
    return (ci, sub, so, partner, name)

inserted_p = 0
inserted_v = 0
errors = 0
seq = 0
for (cat, sub, partner, name), grp in sorted(groups.items(), key=gkey):
    first = grp[0]
    cat_id = cat_map.get(cat)
    if not cat_id:
        print(f'  SKIP no cat: {cat}')
        errors += 1
        continue
    seq += 1
    product_number = f'#P-{seq:03d}'

    product_body = {
        'product_number': product_number,
        'category_id': cat_id,
        'subcategory_id': sub_map.get((cat, sub)),
        'name': name,
        'description': first[ic('description')] or '',
        'base_price': float(first[ic('base_price')]) if first[ic('base_price')] not in ('', None) else 0,
        'price_currency': first[ic('price_currency')] or 'KRW',
        'partner_name': first[ic('partner_name')] or partner,
        'duration_value': int(first[ic('duration_value')]) if first[ic('duration_value')] not in ('', None) else None,
        'duration_unit': first[ic('duration_unit')] or None,
        'has_female_doctor': normalize_bool(first[ic('has_female_doctor')]),
        'has_prayer_room': normalize_bool(first[ic('has_prayer_room')]),
        'dietary_type': normalize_dietary(first[ic('dietary_type')]),
        'location_address': first[ic('location_address')] or None,
        'is_active': normalize_bool(first[ic('is_active')]) if first[ic('is_active')] not in ('', None) else True,
    }
    product_body = {k: v for k, v in product_body.items() if v is not None}

    try:
        result = supa('POST', 'products', body=product_body)
        product_id = result[0]['id']
        inserted_p += 1
    except Exception as e:
        print(f'  FAIL product {product_number}: {e}')
        errors += 1
        continue

    is_solo = len(grp) == 1
    variant_bodies = []
    for i, vrow in enumerate(grp):
        price = vrow[ic('base_price')]
        if price in ('', None):
            price = 0
        raw_label = vrow[ic('variant_label')]
        label = None if is_solo else (raw_label or None)
        variant_bodies.append({
            'product_id': product_id,
            'variant_label': label,
            'base_price': float(price),
            'price_currency': vrow[ic('price_currency')] or 'KRW',
            'sort_order': i,
            'is_active': normalize_bool(vrow[ic('is_active')]) if vrow[ic('is_active')] not in ('', None) else True,
        })
    try:
        supa('POST', 'product_variants', body=variant_bodies)
        inserted_v += len(variant_bodies)
    except Exception as e:
        print(f'  FAIL variants {product_number}: {e}')
        errors += 1

print(f'\nDone. Products: {inserted_p}, Variants: {inserted_v}, Errors: {errors}')
