"""Build products_master_v14.xlsx from v13.

Changes:
  1. K-Beauty subcategory rule: clinic name (Dr.Tune's / Nest Clinic / Selena
     Clinic / Ruby / Dekabi / DIAR) — one Beauty clinic offers many procedures,
     so subcategory-by-clinic gives a useful Beauty filter axis.
  2. K-Medical subcategory rule: medical specialty (Health Check-up / Eye /
     Dental / Korean Medicine / Gynecology). For now only Health Check-up
     items exist; structure ready for the others.
  3. Add ALL DIAR items from the partner price sheet (pdf-to-png 13 pages),
     including à la carte component prices — dermatology procedures stack so
     agents may want to bundle individual treatments.
  4. Add ALL spa items from the source price sheet (sheet "2. K-Beauty"
     non-medical section), including those <1M KRW that were filtered out.

Note on prices:
- DIAR price sheet shows numbers in 만원 (×10,000). VAT 10% NOT included.
- Source spa pricing column is "1,000 KRW" units (×1,000).
"""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook('data/products_master_v13.xlsx')
ws = wb['Products']
header = [c.value for c in ws[1]]
ic = lambda n: header.index(n)

# ── Step 1: Rename K-Beauty subcategory to clinic name ────────────────────
# Existing K-Beauty partner_short values: Dr.Tune's / Nest Clinic / Selena Clinic / Ruby / Dekabi
renamed_beauty = 0
for row in range(2, ws.max_row + 1):
    cat = ws.cell(row=row, column=ic('category') + 1).value
    if cat == 'K-Beauty':
        partner = (ws.cell(row=row, column=ic('partner_short') + 1).value
                   or ws.cell(row=row, column=ic('partner_name') + 1).value
                   or 'Unknown')
        ws.cell(row=row, column=ic('subcategory') + 1, value=partner)
        renamed_beauty += 1
print(f"K-Beauty subcategory renamed: {renamed_beauty}")

# K-Medical: leave as-is (all "Health Check-up"); structure already in place.

# Find next product number + per-cat sort order
last_num = 0
cat_max_sort = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    pn = r[ic('product_number')]
    if pn and isinstance(pn, str) and pn.startswith('#P-'):
        try:
            last_num = max(last_num, int(pn[3:]))
        except ValueError:
            pass
    cat = r[ic('category')]
    so = r[ic('sort_order')]
    if cat and isinstance(so, (int, float)):
        cat_max_sort[cat] = max(cat_max_sort.get(cat, 0), int(so))

def empty_row():
    return {h: '' for h in header}

new_rows = []

# ── Step 3: DIAR full price sheet ──────────────────────────────────────────
# All prices in 만원 → multiply by 10,000 for KRW. VAT 10% excluded.
diar_base = {
    'category': 'K-Beauty',
    'subcategory': 'DIAR',
    'partner_name': 'DIAR Clinic',
    'partner_short': 'DIAR',
    'price_currency': 'KRW',
    'why_recommendation': 'Premium dermatology clinic with comprehensive lifting / stem-cell / signature programs; full procedure menu suitable for layered treatments.',
    'location_address': 'Seoul',
    'has_prayer_room': '', 'dietary_type': 'none',
    'is_active': 'TRUE',
    'source_sheet': 'DIAR/디아르표준수가표.pdf (pdf-to-png)',
    'notes': 'Prices exclude 10% VAT.',
}

# Format: (subcat_section, name, grade, duration_value, duration_unit, price_man, description)
# Where price_man is "만원" units (KRW = price_man * 10_000). None if TBD/quote.
diar_items = [
    # Page 2: Membership
    ('Membership', 'DIAR Membership — Red', 'Standard', None, None, 500, 'Pre-paid membership tier; +5% bonus → effective ₩5,250,000 credit.'),
    ('Membership', 'DIAR Membership — Purple', 'Premium', None, None, 1000, 'Pre-paid membership tier; +10% bonus → effective ₩11,000,000 credit.'),
    ('Membership', 'DIAR Membership — Black', 'VIP Premium', None, None, 2000, 'Pre-paid membership tier; +15% bonus → effective ₩23,000,000 credit.'),

    # Page 3: Kim Kardashian Stem Cell (DIAR MCT)
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Face (1 session)', 'Standard', 60, 'minutes', 140, 'MCT stem cell injection + Estheluxe + Exosome mask.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Face (3 sessions)', 'Premium', 60, 'minutes', 336, 'MCT stem cell injection + Estheluxe + Exosome mask. 3-session course.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Face (5 sessions)', 'VIP Premium', 60, 'minutes', 490, 'MCT stem cell injection + Estheluxe + Exosome mask. 5-session course.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Face + Stem-Cell Fusion (1 session)', 'Premium', 90, 'minutes', 550, 'MCT stem cell injection + MCT triple liquid (exosome / ATP / stem cell) + LDM + Estheluxe + Exosome mask.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Face + Stem-Cell Fusion (3 sessions)', 'VIP Premium', 90, 'minutes', 1320, 'MCT stem cell + MCT triple liquid + LDM + Estheluxe + Exosome mask. 3-session course.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Face + Stem-Cell Fusion (5 sessions)', 'VVIP', 90, 'minutes', 1925, 'MCT stem cell + MCT triple liquid + LDM + Estheluxe + Exosome mask. 5-session course.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Scalp (1 session)', 'Standard', 60, 'minutes', 220, 'Scalp purifying + cleansing care + MCT stem cell hair injection + Estheluxe.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Scalp (3 sessions)', 'Premium', 60, 'minutes', 528, 'Scalp purifying + cleansing care + MCT stem cell hair injection + Estheluxe. 3-session course.'),
    ('Kim Kardashian MCT Stem Cell', 'MCT Stem Cell — Scalp (5 sessions)', 'VIP Premium', 60, 'minutes', 770, 'Scalp purifying + cleansing care + MCT stem cell hair injection + Estheluxe. 5-session course.'),

    # Page 4: 탄력/리프팅 (à la carte component pricing)
    ('Lifting — Component', 'Thermage FLX 300 shots', 'Standard', None, None, 165, 'RF skin tightening, 300 shots.'),
    ('Lifting — Component', 'Thermage FLX 600 shots', 'Standard', None, None, 275, 'RF skin tightening, 600 shots.'),
    ('Lifting — Component', 'Thermage FLX 900 shots', 'Premium', None, None, 440, 'RF skin tightening, 900 shots.'),
    ('Lifting — Component', 'Eye Thermage 225 shots', 'Standard', None, None, 150, 'Periocular RF tightening, 225 shots.'),
    ('Lifting — Component', 'Eye Thermage 450 shots', 'Standard', None, None, 199, 'Periocular RF tightening, 450 shots.'),
    ('Lifting — Component', 'Ultherapy PRIME 100 shots', 'Standard', None, None, 50, 'HIFU lifting, 100 shots.'),
    ('Lifting — Component', 'Ultherapy PRIME 400 shots', 'Standard', None, None, 180, 'HIFU lifting, 400 shots.'),
    ('Lifting — Component', 'Ultherapy PRIME 600 shots', 'Standard', None, None, 260, 'HIFU lifting, 600 shots.'),
    ('Lifting — Component', 'Titanium Lifting 40 kJ', 'Standard', None, None, 60, 'Microfocused ultrasound lifting, 40 kJ.'),
    ('Lifting — Component', 'Titanium Lifting 60 kJ', 'Standard', None, None, 80, 'Microfocused ultrasound lifting, 60 kJ.'),
    ('Lifting — Component', 'ONDA Lifting (Face) 60 kJ', 'Standard', None, None, 60, 'Micro-wave fat reduction & lifting, face, 60 kJ.'),
    ('Lifting — Component', 'ONDA Lifting (Face) 80 kJ', 'Standard', None, None, 80, 'Micro-wave fat reduction & lifting, face, 80 kJ.'),
    ('Lifting — Component', 'ONDA Lifting (Face) 100 kJ', 'Standard', None, None, 100, 'Micro-wave fat reduction & lifting, face, 100 kJ.'),
    ('Lifting — Component', 'ONDA Lifting (Body) 80 kJ', 'Standard', None, None, 96, 'Micro-wave fat reduction, body, 80 kJ.'),
    ('Lifting — Component', 'ONDA Lifting (Body) 100 kJ', 'Standard', None, None, 120, 'Micro-wave fat reduction, body, 100 kJ.'),
    ('Lifting — Component', 'Potenza Lifting N25', 'Standard', None, None, 50, 'RF microneedling lifting, N25 tip.'),
    ('Lifting — Component', 'Potenza Lifting CP25 + Exosome', 'Standard', None, None, 90, 'RF microneedling + exosome serum infusion, CP25 tip.'),
    ('Lifting — Component', '10triple Lifting 200 shots', 'Standard', None, None, 90, '10triple HIFU lifting, 200 shots.'),

    # Page 5: DIAR Tension Solution
    ('DIAR Tension Solution', 'DIAR Tension Solution — Basic', 'Standard', None, None, 180, 'Silhouette Soft × 2 strands + α.'),
    ('DIAR Tension Solution', 'DIAR Tension Solution — Royal', 'Premium', None, None, 300, 'Silhouette Soft × 4 strands + α.'),
    ('DIAR Tension Solution', 'DIAR Tension Solution — Signature', 'VIP Premium', None, None, 420, 'Silhouette Soft × 6 strands + α.'),
    ('DIAR Tension Solution', 'DIAR Tension Solution — Premium', 'VIP Premium', None, None, 500, 'DIAR Premium contouring & tension solution package.'),

    # Page 6: 스킨부스터 (Skin Booster)
    ('Skin Booster', 'Juvelook Volume / Skin 1cc', 'Standard', None, None, 10, 'PDLLA skin booster, 1cc.'),
    ('Skin Booster', 'Radiesse (Collagen Booster) 1syr', 'Standard', None, None, 99, 'CaHA collagen booster, 1 syringe.'),
    ('Skin Booster', 'Juveacell 5cc / 1syr', 'Standard', None, None, 120, 'Salmon DNA polynucleotide booster, 5cc.'),
    ('Skin Booster', 'Juveacell 10cc / 2syr', 'Standard', None, None, 200, 'Salmon DNA polynucleotide booster, 10cc.'),
    ('Skin Booster', 'Juveacell 15cc / 3syr', 'Standard', None, None, 280, 'Salmon DNA polynucleotide booster, 15cc.'),
    ('Skin Booster', 'RE2O 5cc / 1syr', 'Standard', None, None, 100, 'Hydration & elasticity booster, 5cc.'),
    ('Skin Booster', 'RE2O 10cc / 2syr', 'Standard', None, None, 180, 'Hydration & elasticity booster, 10cc.'),
    ('Skin Booster', 'RE2O 15cc / 3syr', 'Standard', None, None, 260, 'Hydration & elasticity booster, 15cc.'),
    ('Skin Booster', 'Gouri (PCL Injection) 6cc / 1syr', 'Standard', None, None, 100, 'PCL polymer collagen booster, 6cc.'),
    ('Skin Booster', 'Gouri (PCL Injection) 12cc / 2syr', 'Standard', None, None, 180, 'PCL polymer collagen booster, 12cc.'),
    ('Skin Booster', 'Gouri (PCL Injection) 18cc / 3syr', 'Standard', None, None, 260, 'PCL polymer collagen booster, 18cc.'),

    # Page 7: 스킨부스터 — imported
    ('Skin Booster', 'Skinvive (Allergan, USA) 1syr', 'Standard', None, None, 42, 'Allergan microdroplet hydrator, 1 syringe.'),
    ('Skin Booster', 'Revive (Merz, Germany) 1syr', 'Standard', None, None, 60, 'Merz collagen-stimulating booster, 1 syringe.'),
    ('Skin Booster', 'Rejuran Healer 1syr', 'Standard', None, None, 32, 'Salmon PN skin healing, 1 syringe.'),
    ('Skin Booster', 'Rejuran Eye 1syr', 'Standard', None, None, 30, 'Periocular skin healing, 1 syringe.'),
    ('DIAR Secret Injection', 'DIAR Secret Injection — Basic', 'Standard', None, None, 80, 'Rejuran Healer 4cc + Dr. Cocktail injection.'),
    ('DIAR Secret Injection', 'DIAR Secret Injection — Premium', 'Premium', None, None, 100, 'Rejuran Healer 4cc + Dr. Cocktail Premium injection.'),

    # Page 8: 필러 (Filler)
    ('Filler', 'Restylane 1cc', 'Standard', None, None, 55, 'Restylane HA filler, 1cc.'),
    ('Filler', 'Juvederm 1cc', 'Standard', None, None, 49, 'Juvederm HA filler, 1cc.'),
    ('Filler', 'Domestic Filler 1cc', 'Standard', None, None, 25, 'Domestic HA filler, 1cc.'),

    # Page 8: 보톡스 (Botox) — by site, by brand (Domestic / Xeomin / Allergan)
    ('Botox', 'Square jaw — Domestic Botox 50u', 'Standard', None, None, 9.9, 'Masseter reduction, domestic toxin.'),
    ('Botox', 'Square jaw — Xeomin (Germany) 50u', 'Standard', None, None, 20, 'Masseter reduction, Xeomin.'),
    ('Botox', 'Square jaw — Allergan (USA) 50u', 'Premium', None, None, 32, 'Masseter reduction, Allergan Botox.'),
    ('Botox', 'Salivary gland / Jaw / Temporal — Domestic', 'Standard', None, None, 25, 'Sialic, jaw, temporal area, domestic toxin.'),
    ('Botox', 'Salivary gland / Jaw / Temporal — Xeomin', 'Standard', None, None, 40, 'Sialic, jaw, temporal area, Xeomin.'),
    ('Botox', 'Salivary gland / Jaw / Temporal — Allergan', 'Premium', None, None, 50, 'Sialic, jaw, temporal area, Allergan.'),
    ('Botox', 'Forehead — Domestic', 'Standard', None, None, 9.9, 'Forehead, domestic.'),
    ('Botox', 'Forehead — Xeomin', 'Standard', None, None, 15, 'Forehead, Xeomin.'),
    ('Botox', 'Forehead — Allergan', 'Premium', None, None, 20, 'Forehead, Allergan.'),
    ('Botox', 'Forehead + Glabella — Domestic', 'Standard', None, None, 15, 'Forehead + glabella, domestic.'),
    ('Botox', 'Forehead + Glabella — Xeomin', 'Standard', None, None, 25, 'Forehead + glabella, Xeomin.'),
    ('Botox', 'Forehead + Glabella — Allergan', 'Premium', None, None, 35, 'Forehead + glabella, Allergan.'),
    ('Botox', 'Barcode (perioral lines) — Domestic', 'Standard', None, None, 12, 'Perioral lines, domestic.'),
    ('Botox', 'Barcode (perioral lines) — Xeomin', 'Standard', None, None, 18, 'Perioral lines, Xeomin.'),
    ('Botox', 'Barcode (perioral lines) — Allergan', 'Premium', None, None, 28, 'Perioral lines, Allergan.'),
    ('Botox', 'Glabella / Crow’s feet / Bunny lines — Domestic', 'Standard', None, None, 9.9, 'Domestic toxin.'),
    ('Botox', 'Glabella / Crow’s feet / Bunny lines — Xeomin', 'Standard', None, None, 15, 'Xeomin.'),
    ('Botox', 'Glabella / Crow’s feet / Bunny lines — Allergan', 'Premium', None, None, 25, 'Allergan.'),
    ('Botox', 'Mouth corners / Chin dimpling — Domestic', 'Standard', None, None, 9.9, 'Domestic toxin.'),
    ('Botox', 'Mouth corners / Chin dimpling — Xeomin', 'Standard', None, None, 18, 'Xeomin.'),
    ('Botox', 'Mouth corners / Chin dimpling — Allergan', 'Premium', None, None, 25, 'Allergan.'),
    ('Botox', 'Eye area / Smile lines / Bunny lines — Domestic', 'Standard', None, None, 15, 'Domestic toxin.'),
    ('Botox', 'Eye area / Smile lines / Bunny lines — Xeomin', 'Standard', None, None, 25, 'Xeomin.'),
    ('Botox', 'Eye area / Smile lines / Bunny lines — Allergan', 'Premium', None, None, 35, 'Allergan.'),
    ('Botox', 'Jawline / Double-chin line — Domestic', 'Standard', None, None, 30, 'Jawline contouring, domestic.'),
    ('Botox', 'Jawline / Double-chin line — Xeomin', 'Standard', None, None, 40, 'Jawline contouring, Xeomin.'),
    ('Botox', 'Jawline / Double-chin line — Allergan', 'Premium', None, None, 50, 'Jawline contouring, Allergan.'),
    ('Botox', 'Dermatoxin (full-face) — Domestic', 'Standard', None, None, 35, 'Microbotox full-face, domestic.'),
    ('Botox', 'Dermatoxin (full-face) — Xeomin', 'Standard', None, None, 45, 'Microbotox full-face, Xeomin.'),
    ('Botox', 'Dermatoxin (full-face) — Allergan', 'Premium', None, None, 55, 'Microbotox full-face, Allergan.'),
    ('Botox', 'Trapezius — Domestic', 'Standard', None, None, 30, 'Trapezius slimming, domestic.'),
    ('Botox', 'Trapezius — Xeomin', 'Standard', None, None, 50, 'Trapezius slimming, Xeomin.'),
    ('Botox', 'Trapezius — Allergan', 'Premium', None, None, 70, 'Trapezius slimming, Allergan.'),
    ('Botox', 'Calf — Domestic', 'Standard', None, None, 55, 'Calf slimming, domestic.'),
    ('Botox', 'Calf — Xeomin', 'Standard', None, None, 88, 'Calf slimming, Xeomin.'),
    ('Botox', 'Calf — Allergan', 'Premium', None, None, 108, 'Calf slimming, Allergan.'),
    ('Botox', 'Underarm hyperhidrosis (100u/200u) — Domestic', 'Standard', None, None, 35, 'Hyperhidrosis 100u/200u, domestic. Two-tier price 35/60.'),
    ('Botox', 'Underarm hyperhidrosis (100u/200u) — Xeomin', 'Standard', None, None, 55, 'Hyperhidrosis 100u/200u, Xeomin. Two-tier price 55/99.'),
    ('Botox', 'Underarm hyperhidrosis (100u/200u) — Allergan', 'Premium', None, None, 70, 'Hyperhidrosis 100u/200u, Allergan. Two-tier price 70/110.'),

    # Page 9: 바디 필러 (Body Filler) — both sides
    ('Body Filler', 'Shoulder filler 20cc / 40cc (both sides)', 'Standard', None, None, 150, 'Both shoulders volumization. Two-tier price 150/210 by volume.'),
    ('Body Filler', 'Clavicle filler (both sides)', 'Standard', None, None, 80, 'Clavicle volumization, both sides.'),
    ('Body Filler', 'Shoulder 40cc + Trapezius Botox', 'Premium', None, None, 230, 'Shoulder filler + trapezius slimming.'),
    ('Body Filler', 'Shoulder 40cc + Clavicle + Trapezius Botox', 'Premium', None, None, 310, 'Shoulder + clavicle filler + trapezius slimming.'),
    ('Body Filler', 'Hip-up / Hip-dip 200cc (both sides)', 'Premium', None, None, 360, 'Hip volumization & contouring, 200cc.'),
    ('Body Filler', 'Hip-up / Hip-dip 400cc (both sides)', 'VIP Premium', None, None, 650, 'Hip volumization & contouring, 400cc.'),
    ('Body Filler', 'Abdomen + Waist 360° ONDA (300 kJ)', 'Premium', None, None, 360, 'Body lifting & contouring around full waistline, 300 kJ ONDA.'),

    # Page 10: DIAR Signature Care
    ('DIAR Signature Care', 'DIAR Signature Skin Care — 2 colors (5 sessions)', 'Premium', None, None, 249, 'Color Laser × 2 + Ion zyme + Estheluxe regeneration. 5-session course.'),
    ('DIAR Signature Care', 'DIAR Signature Skin Care — 3 colors (5 sessions)', 'Premium', None, None, 298, 'Color Laser × 3 + Ion zyme + Estheluxe regeneration. 5-session course.'),
    ('DIAR Signature Care', 'DIAR Signature Skin Care — Combination (5 sessions)', 'VIP Premium', None, None, 399, '[Color Laser × 3 + Half injection + Ion zyme + Estheluxe regen] × 5 + [LaLa Peel + LDM regen + Tension care + WhiteShot (liquid) + Estheluxe] × 5.'),
    ('DIAR One Day All-Kill', 'One Day All-Kill Care — Tentriple+', 'Premium', None, None, 253, 'Tentriple 200 shots + Pico K-Toning + DIAR Secret injection + LDM + DIAR Signature liquid + Halflight injection.'),
    ('DIAR One Day All-Kill', 'One Day All-Kill Care — ONDA+', 'Standard', None, None, 193, 'ONDA 80 kJ + Titanium 40 kJ + Pico K-Toning + Genesis + Ion zyme care + DIAR Signature liquid + Halflight injection.'),

    # Page 11: DIAR Private Program
    ('DIAR Private Program', 'DIAR Premium Lifting — A', 'Premium', None, None, 490, 'Thermage 600 shots + Ultherapy Prime 400 shots + DIAR Secret injection.'),
    ('DIAR Private Program', 'DIAR Premium Lifting — B', 'VIP Premium', None, None, 699, 'Thermage 900 shots + Ultherapy Prime 300 shots + ONDA + DIAR Secret injection.'),
    ('DIAR Private Program', 'DIAR Premium Lifting — C', 'VVIP', None, None, 999, 'Thermage 900 shots + Ultherapy Prime 600 shots + Triple pore care + DIAR Secret injection + RE2O 2syr.'),
    ('DIAR Private Program', 'DIAR Private Stem Cell Care — A (300M cells)', 'Premium', None, None, 120, 'Private stem cell care, 3억 cells.'),
    ('DIAR Private Program', 'DIAR Private Stem Cell Care — B (1.2B cells)', 'VIP Premium', None, None, 400, 'Private stem cell care, 12억 cells.'),
    ('DIAR Private Program', 'DIAR Private Stem Cell Care — C (2.4B cells)', 'VVIP', None, None, 700, 'Private stem cell care, 24억 cells.'),
    ('DIAR Neck Core Care', 'DIAR Neck Core Care — Basic', 'Standard', None, None, 143, 'Tentriple 100 shots + Velotero soft + skin botox.'),
    ('DIAR Neck Core Care', 'DIAR Neck Core Care — Premium', 'Premium', None, None, 271, 'Ultherapy Prime 200 shots + Radiesse + Velotero soft + skin botox.'),

    # Page 12: 여드름/색소/모공/필링 (Acne/Pigment/Pore/Peel)
    ('Acne / Pigment / Pore', 'Potenza Agnes — Full face', 'Standard', None, None, 30, 'RF microneedling for active acne.'),
    ('Acne / Pigment / Pore', 'Potenza Agnes — T-zone', 'Standard', None, None, 20, 'RF microneedling for T-zone acne.'),
    ('Acne / Pigment / Pore', 'Potenza Agnes — Forehead/Jaw', 'Standard', None, None, 15, 'RF microneedling, forehead or jaw.'),
    ('Acne / Pigment / Pore', 'Clarity Acne — Full face', 'Standard', None, None, 30, 'Clarity laser for acne, full face.'),
    ('Acne / Pigment / Pore', 'Clarity Acne — T-zone', 'Standard', None, None, 20, 'Clarity laser for acne, T-zone.'),
    ('Acne / Pigment / Pore', 'Pore Laser — 2 types', 'Standard', None, None, 60, 'Pore tightening, 2-laser combo.'),
    ('Acne / Pigment / Pore', 'Pore Laser — 3 types', 'Standard', None, None, 70, 'Pore tightening, 3-laser combo.'),
    ('Acne / Pigment / Pore', 'Pore Laser — 3 types + Potenza pore boost', 'Premium', None, None, 100, 'Pore tightening 3-laser combo + Potenza pore boost.'),
    ('Acne / Pigment / Pore', 'Pigment Laser — 2 types', 'Standard', None, None, 50, 'Pigment correction, 2-laser combo.'),
    ('Acne / Pigment / Pore', 'Pigment Laser — 3 types', 'Standard', None, None, 70, 'Pigment correction, 3-laser combo.'),
    ('Peeling', 'LaLa Peel — Full face', 'Standard', None, None, 15, 'LaLa peel exfoliation, full face.'),
    ('Peeling', 'LaLa Peel — Face + Neck', 'Standard', None, None, 20, 'LaLa peel exfoliation, face + neck.'),
    ('Peeling', 'Combes’ Peel', 'Standard', None, None, 10, 'Combes deep peeling.'),
    ('Peeling', 'PDT', 'Standard', None, None, 10, 'Photodynamic therapy.'),
    ('Peeling', 'LDM', 'Standard', None, None, 10, 'Local dynamic micromassage soothing care.'),
    ('Peeling', 'SONO', 'Standard', None, None, 5, 'Sono cleanser ultrasonic care.'),

    # Page 13: 페이스·바디 제모 (Hair Removal) — both sides combined; 1회 / 5회
    ('Hair Removal', 'Full face (incl. hairline) — 1 session', 'Standard', None, None, 20, 'Laser hair removal, full face including hairline.'),
    ('Hair Removal', 'Full face (incl. hairline) — 5 sessions', 'Standard', None, None, 60, 'Laser hair removal, full face including hairline. 5-session course.'),
    ('Hair Removal', 'Philtrum / Chin — 1 session', 'Standard', None, None, 10, 'Philtrum or chin, single session.'),
    ('Hair Removal', 'Philtrum / Chin — 5 sessions', 'Standard', None, None, 30, 'Philtrum or chin, 5-session course.'),
    ('Hair Removal', 'Underarms — 1 session', 'Standard', None, None, 8, 'Underarm laser hair removal, single session.'),
    ('Hair Removal', 'Underarms — 5 sessions', 'Standard', None, None, 24, 'Underarm laser hair removal, 5-session course.'),
    ('Hair Removal', 'Lower arm — 1 session', 'Standard', None, None, 15, 'Lower arm hair removal.'),
    ('Hair Removal', 'Lower arm — 5 sessions', 'Standard', None, None, 45, 'Lower arm hair removal, 5-session course.'),
    ('Hair Removal', 'Full arm — 1 session', 'Standard', None, None, 25, 'Full arm hair removal.'),
    ('Hair Removal', 'Full arm — 5 sessions', 'Standard', None, None, 75, 'Full arm hair removal, 5-session course.'),
    ('Hair Removal', 'Calf (below knee) — 1 session', 'Standard', None, None, 25, 'Calf hair removal.'),
    ('Hair Removal', 'Calf (below knee) — 5 sessions', 'Standard', None, None, 75, 'Calf hair removal, 5-session course.'),
    ('Hair Removal', 'Full leg — 1 session', 'Standard', None, None, 50, 'Full leg hair removal.'),
    ('Hair Removal', 'Full leg — 5 sessions', 'Standard', None, None, 150, 'Full leg hair removal, 5-session course.'),
    ('Hair Removal', 'Belly button line — 1 session', 'Standard', None, None, 10, 'Belly button line hair removal.'),
    ('Hair Removal', 'Belly button line — 5 sessions', 'Standard', None, None, 30, 'Belly button line hair removal, 5-session course.'),
    ('Hair Removal', 'Belly button to lower abdomen — 1 session', 'Standard', None, None, 20, 'Lower abdomen hair removal.'),
    ('Hair Removal', 'Belly button to lower abdomen — 5 sessions', 'Standard', None, None, 60, 'Lower abdomen hair removal, 5-session course.'),
    ('Hair Removal', 'Front chest — 1 session', 'Standard', None, None, 25, 'Front chest hair removal.'),
    ('Hair Removal', 'Front chest — 5 sessions', 'Standard', None, None, 75, 'Front chest hair removal, 5-session course.'),
    ('Hair Removal', 'Front belly (bra line ~ pubic line) — 1 session', 'Standard', None, None, 25, 'Belly hair removal between bra line and pubic line.'),
    ('Hair Removal', 'Front belly (bra line ~ pubic line) — 5 sessions', 'Standard', None, None, 75, 'Belly hair removal, 5-session course.'),
    ('Hair Removal', 'Back / Lower back — 1 session', 'Standard', None, None, 50, 'Back & lower back hair removal.'),
    ('Hair Removal', 'Back / Lower back — 5 sessions', 'Standard', None, None, 150, 'Back & lower back hair removal, 5-session course.'),
    ('Hair Removal', 'Bikini — 1 session', 'Standard', None, None, 15, 'Bikini line hair removal.'),
    ('Hair Removal', 'Bikini — 5 sessions', 'Standard', None, None, 45, 'Bikini line hair removal, 5-session course.'),
    ('Hair Removal', 'Brazilian — 1 session', 'Standard', None, None, 25, 'Brazilian hair removal.'),
    ('Hair Removal', 'Brazilian — 5 sessions', 'Standard', None, None, 75, 'Brazilian hair removal, 5-session course.'),
    ('Hair Removal', 'Buttocks — 1 session', 'Standard', None, None, 20, 'Buttocks hair removal.'),
    ('Hair Removal', 'Buttocks — 5 sessions', 'Standard', None, None, 60, 'Buttocks hair removal, 5-session course.'),
]

sort_order = cat_max_sort.get('K-Beauty', 0)
for sub_section, name, grade, dv, du, price_man, desc in diar_items:
    last_num += 1
    sort_order += 1
    r = empty_row()
    r.update(diar_base)
    r['product_number'] = f'#P-{last_num:03d}'
    r['name'] = name
    r['grade'] = grade
    r['base_price'] = int(price_man * 10_000) if price_man is not None else ''
    r['description'] = desc
    r['duration_value'] = dv if dv is not None else ''
    r['duration_unit'] = du if du is not None else ''
    r['sort_order'] = sort_order
    # Subcategory still "DIAR" — sub_section is informational (could go in notes)
    r['notes'] = sub_section
    new_rows.append(r)
print(f"DIAR items added: {len(diar_items)}")

# ── Step 4: Add ALL spa items from price sheet "2. K-Beauty" non-medical ──
# Including those <1M KRW that were filtered out before. Place under K-Wellness > Spa.
# Source quotation column unit = "1,000 KRW" → KRW = price_k * 1000.
spa_items = [
    # SPA the belec — Incheon Songdo
    {'partner_name': 'SPA the belec', 'partner_short': 'SPA the belec', 'name': 'Mix (Silver)',
     'grade': 'Standard', 'price_k': 297, 'duration': 100,
     'desc': 'Basic Face + Body efficient care.',
     'addr': 'Incheon Songdo', 'contact_email': '', 'info_url': 'instagram: @_thebelec'},
    {'partner_name': 'SPA the belec', 'partner_short': 'SPA the belec', 'name': 'Spa (Silver)',
     'grade': 'Standard', 'price_k': 363, 'duration': 120,
     'desc': 'Head to Foot Full Line — body care managed with spa-specific protocols.',
     'addr': 'Incheon Songdo', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'SPA the belec', 'partner_short': 'SPA the belec', 'name': 'Medi Course (Diamond Premium) — Belec Line 2/3',
     'grade': 'Premium', 'price_k': 425, 'duration': 120,
     'desc': 'Belec Line 2 or 3-Body Care — restore your body’s natural balance.',
     'addr': 'Incheon Songdo', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'SPA the belec', 'partner_short': 'SPA the belec', 'name': 'Medi Course (Diamond Premium) — Medi Slimming 2/3',
     'grade': 'Premium', 'price_k': 473, 'duration': 120,
     'desc': 'Medi Slimming 2 or 3-Body Care (body slimming care).',
     'addr': 'Incheon Songdo', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'SPA the belec', 'partner_short': 'SPA the belec', 'name': 'Special Course (Gold)',
     'grade': 'Premium', 'price_k': 393, 'duration': 120,
     'desc': 'Elimination of active oxygen and waste products from the body.',
     'addr': 'Incheon Songdo', 'contact_email': '', 'info_url': ''},

    # Retreat SIGNIEL SPA — Songpa-gu
    {'partner_name': 'Retreat SIGNIEL SPA', 'partner_short': 'Signiel Spa', 'name': 'Contour Whole Body Program (Gold)',
     'grade': 'Premium', 'price_k': 670, 'duration': 150,
     'desc': 'Full-body detox combined with a V-line facial.',
     'addr': 'Gangnam Songpa-gu, Seoul', 'contact_email': '', 'info_url': 'https://www.lottehotel.com/seoul-signiel/en/facilities/spa.html'},
    {'partner_name': 'Retreat SIGNIEL SPA', 'partner_short': 'Signiel Spa', 'name': 'Diamond Détente (Gold)',
     'grade': 'Premium', 'price_k': 680, 'duration': 150,
     'desc': 'Luxurious full-body treatment with diamond powder massage.',
     'addr': 'Gangnam Songpa-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Retreat SIGNIEL SPA', 'partner_short': 'Signiel Spa', 'name': 'Timeless Away Journey (Diamond Premium)',
     'grade': 'VIP Premium', 'price_k': 750, 'duration': 180,
     'desc': 'Restorative body ritual with coffee scrub and baobab oil.',
     'addr': 'Gangnam Songpa-gu, Seoul', 'contact_email': '', 'info_url': ''},
    # Chocolate Whole Body — already in v13, skip duplicate

    # Four Seasons Hotel SPA — Jongno-gu
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Serenity with Thermo-Stones (Gold)',
     'grade': 'Premium', 'price_k': 340, 'duration': 90,
     'desc': 'Healing treatment relieving body tension with warm thermo-stones.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': 'https://www.fourseasons.com/seoul/spa/'},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Bespoke Therapeutic (Silver)',
     'grade': 'Standard', 'price_k': 320, 'duration': 90,
     'desc': 'Luxurious aromatherapy massage tailored to your needs.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Revitalize with K-Wellness (Silver)',
     'grade': 'Standard', 'price_k': 260, 'duration': 60,
     'desc': 'Korean-inspired rejuvenating treatment.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Mandarin Deep Tissue (Gold)',
     'grade': 'Premium', 'price_k': 330, 'duration': 90,
     'desc': 'Targeted remedial deep tissue treatment for stiff areas.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Seoul’s Luxurious Renewal Journey (Diamond Premium)',
     'grade': 'VIP Premium', 'price_k': 530, 'duration': 150,
     'desc': 'Modern take on Korea’s historic treatment traditions.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Romantic Retreat (Gold)',
     'grade': 'Premium', 'price_k': 530, 'duration': 150,
     'desc': 'Light-filled spa suite couple’s warm immersion.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Wellness Balancing Treatment — Premium Upgrade (Silver)',
     'grade': 'Standard', 'price_k': 530, 'duration': 120,
     'desc': 'Warm volcanic stone treatment targeting weariness.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Revolutionize with Second Skin (Diamond Premium)',
     'grade': 'VIP Premium', 'price_k': 560, 'duration': 75,
     'desc': 'Revolutionary "Second Skin" facial treatment.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'Bespoke Bliss for Radiant Skin (Gold)',
     'grade': 'Premium', 'price_k': 450, 'duration': 90,
     'desc': 'Biologique Recherche methodology custom-tailored facial.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},
    {'partner_name': 'Four Seasons Hotel SPA', 'partner_short': 'Four Seasons Spa', 'name': 'MediSpa Pro (Silver)',
     'grade': 'Standard', 'price_k': 410, 'duration': 90,
     'desc': '90-minute facial featuring Cerage technology.',
     'addr': 'Jongno-gu, Seoul', 'contact_email': '', 'info_url': ''},

    # Eco Jardin — scalp
    {'partner_name': 'ECO JARDIN', 'partner_short': 'Eco Jardin', 'name': 'Premium Scalp Treatment',
     'grade': 'Standard', 'price_k': 250, 'duration': '',
     'desc': 'Premium scalp care program — purifying, exfoliation, and revitalization.',
     'addr': 'Seoul', 'contact_email': 'k.official@ecojardin.co.kr', 'info_url': 'https://ecojardin.co.kr/kr/spa'},
]

spa_max = cat_max_sort.get('K-Wellness', 0)
for it in spa_items:
    last_num += 1
    spa_max += 1
    r = empty_row()
    r.update({
        'product_number': f'#P-{last_num:03d}',
        'category': 'K-Wellness',
        'subcategory': 'Spa',
        'partner_name': it['partner_name'],
        'partner_short': it['partner_short'],
        'name': it['name'],
        'grade': it['grade'],
        'base_price': it['price_k'] * 1000,
        'price_currency': 'KRW',
        'duration_value': it['duration'] if it['duration'] != '' else '',
        'duration_unit': 'minutes' if it['duration'] != '' else '',
        'description': it['desc'],
        'why_recommendation': 'Premium spa partner — full menu suitable for relaxation programming alongside Beauty/Wellness flow.',
        'location_address': it['addr'],
        'contact_email': it['contact_email'],
        'info_url': it['info_url'],
        'has_prayer_room': '', 'dietary_type': 'none',
        'is_active': 'TRUE',
        'sort_order': spa_max,
        'source_sheet': 'Internal_Price sheet — 2. K-Beauty (non-medical)',
    })
    new_rows.append(r)
print(f"Spa items added: {len(spa_items)}")

# Append all new rows
start_row = ws.max_row + 1
for i, row in enumerate(new_rows):
    for col_idx, h in enumerate(header, 1):
        ws.cell(row=start_row + i, column=col_idx, value=row.get(h, ''))

# Recompute Summary
counts = Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    v = r[ic('category')]
    if v:
        counts[v] += 1
summary = wb['Summary']
for row in summary.iter_rows(min_row=2):
    for c in row:
        c.value = None
ORDER = ['K-Medical', 'K-Beauty', 'K-Wellness', 'K-Education', 'K-Starcation', 'Subpackage']
total = 0
ri = 2
for cat in ORDER:
    n = counts.get(cat, 0)
    summary.cell(row=ri, column=1, value=cat)
    summary.cell(row=ri, column=2, value=n)
    total += n
    ri += 1
summary.cell(row=ri, column=1, value='TOTAL')
summary.cell(row=ri, column=2, value=total)

wb.save('data/products_master_v14.xlsx')
print(f"\nv14 totals:")
for cat in ORDER:
    print(f"  {cat}: {counts.get(cat, 0)}")
print(f"  TOTAL: {total}")
