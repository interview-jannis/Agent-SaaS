"""Build products_master_v15.xlsx from v14.

Changes:
  1. Add `variant_label` column right after `grade`. Same base product with
     different session counts / shot counts / quantities now shares the same
     `name`, distinguished by variant_label. Designed so the SaaS UI can
     group by (partner_short, name) and offer a variant picker on the product
     card / detail modal.
  2. Sort all Product rows by column B (category) for browsability, with
     sort_order as a tie-breaker within each category.
"""
import re
from openpyxl import load_workbook

wb = load_workbook('data/products_master_v14.xlsx')
ws = wb['Products']
header = [c.value for c in ws[1]]

# ── Step 1: Insert variant_label column right after grade ────────────────
grade_idx = header.index('grade') + 1  # 1-indexed
variant_col_pos = grade_idx + 1
ws.insert_cols(variant_col_pos)
ws.cell(row=1, column=variant_col_pos, value='variant_label')
header = [c.value for c in ws[1]]
ic = lambda n: header.index(n)

# ── Step 2: Parse variants from name on relevant rows ────────────────────
# Patterns to extract; each returns (base_name, variant_label) or None.
PATTERNS = [
    # "(1 session)" / "(3 sessions)" / "(5 sessions)" suffix
    (re.compile(r'^(.*)\s*\((\d+\s*sessions?)\)\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "— 1 session" / "— 5 sessions" suffix
    (re.compile(r'^(.*?)\s+[—-]\s+(\d+\s*sessions?)\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "300 shots" trailing — applies to Thermage/Ultherapy/etc
    (re.compile(r'^(.*?)\s+(\d+\s*shots)\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "100 kJ" / "60 kJ"
    (re.compile(r'^(.*?)\s+(\d+\s*kJ)\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "5cc / 1syr" / "10cc / 2syr"
    (re.compile(r'^(.*?)\s+(\d+\s*cc\s*/\s*\d+\s*syr)\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "1cc" / "1syr" trailing
    (re.compile(r'^(.*?)\s+(\d+\s*(?:cc|syr))\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "— Basic" / "— Royal" / "— Signature" / "— Premium" / "— A" / "— B" / "— C" suffix (DIAR Tension/Premium Lifting)
    (re.compile(r'^(.*?)\s+[—-]\s+(Basic|Royal|Signature|Premium|A|B|C|Red|Purple|Black)\s*$'), lambda m: (m.group(1).strip(), m.group(2).strip())),
    # "— Domestic" / "— Xeomin (Germany)" / "— Allergan (USA)" — Botox brand variant
    (re.compile(r'^(.*?)\s+[—-]\s+(Domestic|Xeomin(?:\s*\([^)]+\))?|Allergan(?:\s*\([^)]+\))?)\s*(?:\d*u?)?\s*$'),
     lambda m: (m.group(1).strip(), m.group(2).strip())),
    # Generic suffix in parens at end if name still has variant-like info — skip (too risky)
]

variants_extracted = 0
for row in range(2, ws.max_row + 1):
    name_cell = ws.cell(row=row, column=ic('name') + 1)
    nm = name_cell.value
    if not nm or not isinstance(nm, str):
        continue
    for pat, extractor in PATTERNS:
        m = pat.match(nm)
        if m:
            base, variant = extractor(m)
            if base and variant:
                name_cell.value = base
                ws.cell(row=row, column=ic('variant_label') + 1, value=variant)
                variants_extracted += 1
                break

print(f"Variant labels extracted: {variants_extracted}")

# ── Step 3: Sort Products by category (B) then sort_order ────────────────
# Read all data rows
data_rows = []
for row_idx in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
    data_rows.append(row_vals)

# Define category order; unknowns go to the end alphabetically
CAT_ORDER = ['K-Medical', 'K-Beauty', 'K-Wellness', 'K-Education', 'K-Starcation', 'Subpackage']
def cat_key(row):
    cat = row[ic('category')] or ''
    try:
        return (CAT_ORDER.index(cat), 0, 0)
    except ValueError:
        return (len(CAT_ORDER), cat, 0)

def full_key(row):
    cat = row[ic('category')] or ''
    sub = row[ic('subcategory')] or ''
    so = row[ic('sort_order')]
    so = so if isinstance(so, (int, float)) else 999999
    try:
        ci = CAT_ORDER.index(cat)
    except ValueError:
        ci = len(CAT_ORDER)
    return (ci, sub, so)

data_rows.sort(key=full_key)

# Re-number product_number to keep them increasing in display order
pn_idx = ic('product_number')
for i, row in enumerate(data_rows, start=1):
    row[pn_idx] = f'#P-{i:03d}'

# Write back
for row_idx, row in enumerate(data_rows, start=2):
    for col_idx, val in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

print(f"Sorted {len(data_rows)} rows; product numbers re-issued.")

wb.save('data/products_master_v15.xlsx')
print("Saved data/products_master_v15.xlsx")
